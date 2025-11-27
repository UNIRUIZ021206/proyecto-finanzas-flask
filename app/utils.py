# app/utils.py
import math
import os
import bcrypt
try:
    import google.generativeai as genai
except ImportError:
    class MockGenAI:
        class GenerativeModel:
            def __init__(self, model_name): pass
            def generate_content(self, prompt): return MockResponse()
    class MockResponse:
        text = "Análisis IA no disponible (librería faltante)."
    genai = MockGenAI()
from markdown import markdown
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from sqlalchemy import text
from functools import wraps
from flask import flash, redirect, url_for
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Importamos el engine compartido y la clave de API desde extensions
from .extensions import engine, GEMINI_API_KEY

# --- Filtro personalizado ---
# Nota: El decorador @app.template_filter se aplica en __init__.py
def is_inf(value):
    """Filtro para verificar si un valor es infinito"""
    try:
        return math.isinf(float(value))
    except (ValueError, TypeError):
        return False

# --- Funciones auxiliares para trabajar con Roles ---

def get_rol_id_by_name(nombre_rol):
    """Obtiene el Id_Rol desde la tabla Roles."""
    try:
        with engine.connect() as conn:
            query = text("SELECT Id_Rol FROM Roles WHERE Nombre = :nombre AND Estado = 1")
            result = conn.execute(query, {"nombre": nombre_rol}).fetchone()
            if result:
                return result[0]
    except Exception as e:
        print(f"Error al obtener Id_Rol para {nombre_rol}: {e}")
    return None

def get_rol_name_by_id(id_rol):
    """Obtiene el nombre del rol desde la tabla Roles."""
    try:
        with engine.connect() as conn:
            query = text("SELECT Nombre FROM Roles WHERE Id_Rol = :id_rol AND Estado = 1")
            result = conn.execute(query, {"id_rol": id_rol}).fetchone()
            if result:
                return result[0]
    except Exception as e:
        print(f"Error al obtener nombre de rol para Id_Rol {id_rol}: {e}")
    return None

def is_user_role(user_id, nombre_rol):
    """Verifica si un usuario tiene un rol específico."""
    try:
        with engine.connect() as conn:
            query = text("""
                SELECT COUNT(*) FROM Usuarios u
                INNER JOIN Roles r ON u.Id_Rol = r.Id_Rol
                WHERE u.Id_Usuario = :user_id AND r.Nombre = :nombre_rol AND u.Estado = 1 AND r.Estado = 1
            """)
            result = conn.execute(query, {"user_id": user_id, "nombre_rol": nombre_rol}).fetchone()
            return result[0] > 0 if result else False
    except Exception as e:
        print(f"Error al verificar rol {nombre_rol} para usuario {user_id}: {e}")
    return False

def is_admin(user_id):
    """Verifica si un usuario es administrador."""
    admin_roles = [
        'Administrador', 'administrador', 'Admin', 'admin',
        'Super Administrador', 'super administrador', 'SuperAdministrador', 'superadministrador',
        'SA', 'sa', 'Super Admin', 'super admin', 'SuperAdmin', 'superadmin'
    ]
    for admin_name in admin_roles:
        if is_user_role(user_id, admin_name):
            return True
    
    try:
        with engine.connect() as conn:
            query = text("""
                SELECT r.Nombre FROM Usuarios u
                INNER JOIN Roles r ON u.Id_Rol = r.Id_Rol
                WHERE u.Id_Usuario = :user_id AND u.Estado = 1 AND r.Estado = 1
            """)
            result = conn.execute(query, {"user_id": user_id}).fetchone()
            if result:
                rol_nombre = result[0].lower()
                if 'admin' in rol_nombre or 'sa' in rol_nombre:
                    return True
    except Exception as e:
        print(f"Error al verificar si es admin para usuario {user_id}: {e}")
    
    return False

def is_super_admin(user_id):
    """Verifica si un usuario es Super Administrador (SA)."""
    super_admin_names = [
        'Super Administrador', 'super administrador', 'SuperAdministrador', 'superadministrador',
        'SA', 'sa', 'Super Admin', 'super admin', 'SuperAdmin', 'superadmin'
    ]
    for sa_name in super_admin_names:
        if is_user_role(user_id, sa_name):
            return True
    return False

def is_user_inf(user_id):
    """Verifica si un usuario es INF."""
    inf_roles = ['INF', 'inf', 'Inf']
    for inf_name in inf_roles:
        if is_user_role(user_id, inf_name):
            return True
    return False

# --- Decoradores personalizados ---

def admin_required(f):
    """Decorador que verifica si el usuario actual es administrador."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Debes iniciar sesión primero.', 'warning')
            return redirect(url_for('auth.login')) 
        
        if not is_admin(current_user.id):
            flash('No tienes permisos para acceder a esta página.', 'error')
            return redirect(url_for('main.index'))
        
        return f(*args, **kwargs)
    return decorated_function

# --- Funciones para obtener reportes financieros ---

def get_financial_reports(anio_seleccionado):
    """
    Obtiene los datos de Balance General y Estado de Resultados para un año específico,
    con totales por subtipo.
    """
    try:
        with engine.connect() as conn:
            query = text("""
                SELECT
                    c.CuentaID AS cuenta_id,
                    c.NombreCuenta AS cuenta_nombre,
                    c.TipoCuenta AS tipo,
                    c.SubTipoCuenta AS subtipo,
                    COALESCE(s.Monto, 0) AS monto_actual
                FROM
                    CatalogoCuentas c
                INNER JOIN
                    Periodo p ON p.Anio = :anio
                LEFT JOIN
                    SaldoCuenta s ON s.CuentaID = c.CuentaID AND s.PeriodoID = p.PeriodoID
                ORDER BY
                    c.TipoCuenta, c.SubTipoCuenta, c.NombreCuenta
            """)
            
            resultados = conn.execute(query, {"anio": anio_seleccionado}).fetchall()
            
            if not resultados:
                return None 

            # Estructura de datos
            report_data = {
                'Activo': defaultdict(list),
                'Pasivo': defaultdict(list),
                'Patrimonio': defaultdict(list),
                'Ingreso': defaultdict(list),
                'Costo': defaultdict(list),
                'Gasto': defaultdict(list),
                'Totales': defaultdict(float)
            }

            for i, row in enumerate(resultados):
                try:
                    cuenta = {'id': row[0], 'nombre': row[1], 'monto': 0.0} 
                    tipo = str(row[2]).strip() if row[2] else None
                    subtipo = str(row[3]).strip() if row[3] else None
                    monto_actual = float(row[4]) if row[4] is not None else 0.0
                    
                    # Debug print to see what we are getting
                    # print(f"DEBUG: Cuenta: {row[1]}, Tipo: '{tipo}', Subtipo: '{subtipo}', Monto: {monto_actual}")

                    # Validar que el tipo existe en report_data
                    if not tipo:
                        print(f"Tipo nulo para cuenta: {row[1]}")
                        continue
                        
                    # Normalize type to match keys if needed (simple capitalization)
                    # This handles cases like 'pasivo' vs 'Pasivo'
                    tipo_normalized = tipo.title()
                    
                    if tipo_normalized not in report_data:
                        # Try to map common variations just in case
                        if 'Pasivo' in tipo_normalized:
                            tipo_normalized = 'Pasivo'
                        elif 'Patrimonio' in tipo_normalized or 'Capital' in tipo_normalized:
                            tipo_normalized = 'Patrimonio'
                        elif 'Activo' in tipo_normalized:
                            tipo_normalized = 'Activo'
                        elif 'Ingreso' in tipo_normalized:
                            tipo_normalized = 'Ingreso'
                        elif 'Costo' in tipo_normalized:
                            tipo_normalized = 'Costo'
                        elif 'Gasto' in tipo_normalized:
                            tipo_normalized = 'Gasto'
                    
                    if tipo_normalized not in report_data:
                        print(f"Tipo '{tipo}' (normalizado: '{tipo_normalized}') no válido o no encontrado en report_data. Saltando cuenta: {row[1]}")
                        continue
                    
                    # Use the normalized type
                    tipo = tipo_normalized
                    
                    # Si la cuenta contiene "depreciación" o "deprecioacion" en el nombre, hacer el monto negativo
                    nombre_cuenta_lower = str(row[1]).lower() if row[1] else ''
                    if 'depreciaci' in nombre_cuenta_lower or 'deprecioaci' in nombre_cuenta_lower:
                        # Si el monto ya es negativo, mantenerlo negativo; si es positivo, hacerlo negativo
                        if monto_actual > 0:
                            monto_actual = -abs(monto_actual)
                        elif monto_actual == 0:
                            monto_actual = 0.0
                        # Si ya es negativo, mantenerlo así
                    
                    cuenta['monto'] = monto_actual
                    if subtipo:
                        report_data[tipo][subtipo].append(cuenta)
                    report_data['Totales'][tipo] += monto_actual
                    if subtipo:
                        report_data['Totales'][subtipo] += monto_actual
                except Exception as e:
                    print(f"Error procesando fila {i} en get_financial_reports: {e}")
                    print(f"Datos de la fila: {row}")
                    import traceback
                    traceback.print_exc()
                    continue

            # Calcular Totales Principales
            report_data['Totales']['Total Activo'] = report_data['Totales']['Activo']
            report_data['Totales']['Total Pasivo'] = report_data['Totales']['Pasivo']
            report_data['Totales']['Total Patrimonio'] = report_data['Totales']['Patrimonio']
            report_data['Totales']['Total Pasivo y Patrimonio'] = report_data['Totales']['Pasivo'] + report_data['Totales']['Patrimonio']
            
            # Calcular Utilidades
            total_ingresos = report_data['Totales']['Ingreso']
            total_costos = report_data['Totales']['Costo']
            utilidad_bruta = total_ingresos - total_costos
            
            report_data['Totales']['Utilidad Bruta'] = utilidad_bruta
            
            # Calcular Utilidad Operativa (Utilidad Bruta - Gastos Operativos)
            gastos_operativos = report_data['Totales'].get('Gasto Operativo', 0.0)
            utilidad_operativa = utilidad_bruta - gastos_operativos
            report_data['Totales']['Utilidad Operativa'] = utilidad_operativa
            
            utilidad_neta = utilidad_bruta - report_data['Totales']['Gasto']
            report_data['Totales']['Utilidad Neta'] = utilidad_neta
            
            return report_data
            
    except Exception as e:
        print(f"Error EXCEPCIÓN en get_financial_reports: {e}")
        return None 

def calcular_analisis_horizontal(report_data_base, report_data_analisis):
    """Calcula el análisis horizontal comparando dos períodos."""
    analisis = {
        'Activo': defaultdict(list),
        'Pasivo': defaultdict(list),
        'Patrimonio': defaultdict(list),
        'Ingreso': defaultdict(list),
        'Costo': defaultdict(list),
        'Gasto': defaultdict(list),
        'Totales': {}
    }
    
    # Comparar cuentas para cada tipo y subtipo
    for tipo in ['Activo', 'Pasivo', 'Patrimonio', 'Ingreso', 'Costo', 'Gasto']:
        for subtipo in set(list(report_data_base[tipo].keys()) + list(report_data_analisis[tipo].keys())):
            cuentas_base = {c['id']: c for c in report_data_base[tipo].get(subtipo, [])}
            cuentas_analisis = {c['id']: c for c in report_data_analisis[tipo].get(subtipo, [])}
            
            todas_cuentas_ids = set(cuentas_base.keys()) | set(cuentas_analisis.keys())
            
            for cuenta_id in todas_cuentas_ids:
                cuenta_base_obj = cuentas_base.get(cuenta_id, {'nombre': '', 'monto': 0.0})
                cuenta_analisis_obj = cuentas_analisis.get(cuenta_id, {'nombre': cuenta_base_obj['nombre'], 'monto': 0.0})
                
                monto_base = cuenta_base_obj['monto']
                monto_analisis = cuenta_analisis_obj['monto']
                absoluto = monto_analisis - monto_base
                
                if monto_base != 0:
                    relativo = ((monto_analisis / monto_base) - 1) * 100
                else:
                    if monto_analisis != 0:
                        relativo = float('inf')
                    else:
                        relativo = 0.0
                
                # Determinar color
                if relativo < 0:
                    color_clase = 'valor-negativo'
                elif relativo > 0:
                    color_clase = 'valor-positivo'
                else:
                    color_clase = 'valor-cero'
                
                analisis[tipo][subtipo].append({
                    'id': cuenta_id,
                    'nombre': cuenta_analisis_obj['nombre'] or cuenta_base_obj['nombre'],
                    'monto_base': monto_base,
                    'monto_analisis': monto_analisis,
                    'absoluto': absoluto,
                    'relativo': relativo,
                    'color_clase': color_clase
                })
    
    # Calcular totales para tipos principales
    tipos_principales = ['Total Activo', 'Total Pasivo', 'Total Patrimonio', 'Total Pasivo y Patrimonio',
                    'Ingreso', 'Costo', 'Gasto', 'Utilidad Bruta', 'Utilidad Neta']
    
    for key in tipos_principales:
        if key in report_data_base['Totales'] and key in report_data_analisis['Totales']:
            monto_base = report_data_base['Totales'][key]
            monto_analisis = report_data_analisis['Totales'][key]
            absoluto = monto_analisis - monto_base
            
            if monto_base != 0:
                relativo = ((monto_analisis / monto_base) - 1) * 100
            else:
                if monto_analisis != 0:
                    relativo = float('inf')
                else:
                    relativo = 0.0
            
            if relativo < 0:
                color_clase = 'valor-negativo'
            elif relativo > 0:
                color_clase = 'valor-positivo'
            else:
                color_clase = 'valor-cero'
            
            analisis['Totales'][key] = {
                'base': monto_base,
                'analisis': monto_analisis,
                'absoluto': absoluto,
                'relativo': relativo,
                'color_clase': color_clase
            }
        elif key in report_data_base['Totales']:
            # Solo existe en base
            analisis['Totales'][key] = {
                'base': report_data_base['Totales'][key],
                'analisis': 0.0,
                'absoluto': -report_data_base['Totales'][key],
                'relativo': -100.0,
                'color_clase': 'valor-negativo'
            }
        elif key in report_data_analisis['Totales']:
            # Solo existe en análisis
            analisis['Totales'][key] = {
                'base': 0.0,
                'analisis': report_data_analisis['Totales'][key],
                'absoluto': report_data_analisis['Totales'][key],
                'relativo': float('inf'),
                'color_clase': 'valor-positivo'
            }
    
    # Totales especiales para Pasivo y Patrimonio
    if 'Pasivo' in report_data_base['Totales'] and 'Pasivo' in report_data_analisis['Totales']:
        analisis['Totales']['Pasivo'] = {
            'base': report_data_base['Totales']['Pasivo'],
            'analisis': report_data_analisis['Totales']['Pasivo'],
            'absoluto': report_data_analisis['Totales']['Pasivo'] - report_data_base['Totales']['Pasivo'],
            'relativo': ((report_data_analisis['Totales']['Pasivo'] / report_data_base['Totales']['Pasivo']) - 1) * 100 if report_data_base['Totales']['Pasivo'] != 0 else 0.0,
            'color_clase': 'valor-positivo' if report_data_analisis['Totales']['Pasivo'] > report_data_base['Totales']['Pasivo'] else ('valor-negativo' if report_data_analisis['Totales']['Pasivo'] < report_data_base['Totales']['Pasivo'] else 'valor-cero')
        }
    
    if 'Patrimonio' in report_data_base['Totales'] and 'Patrimonio' in report_data_analisis['Totales']:
        analisis['Totales']['Patrimonio'] = {
            'base': report_data_base['Totales']['Patrimonio'],
            'analisis': report_data_analisis['Totales']['Patrimonio'],
            'absoluto': report_data_analisis['Totales']['Patrimonio'] - report_data_base['Totales']['Patrimonio'],
            'relativo': ((report_data_analisis['Totales']['Patrimonio'] / report_data_base['Totales']['Patrimonio']) - 1) * 100 if report_data_base['Totales']['Patrimonio'] != 0 else 0.0,
            'color_clase': 'valor-positivo' if report_data_analisis['Totales']['Patrimonio'] > report_data_base['Totales']['Patrimonio'] else ('valor-negativo' if report_data_analisis['Totales']['Patrimonio'] < report_data_base['Totales']['Patrimonio'] else 'valor-cero')
        }
        
    return analisis

def calcular_ratios_financieros(report_data, report_data_anterior=None):
    """
    Calcula los ratios financieros principales.
    report_data: Datos del año actual
    report_data_anterior: Datos del año anterior (opcional, para promedios)
    """
    ratios = {
        'Liquidez': {},
        'Actividades': {},
        'Endeudamiento': {},
        'Rentabilidad': {}
    }
    totales = report_data['Totales']
    
    # Extraer valores necesarios
    activo_corriente = sum([c['monto'] for c in report_data['Activo'].get('Activo Corriente', [])])
    activo_no_corriente = sum([c['monto'] for c in report_data['Activo'].get('Activo No Corriente', [])])
    total_activo = totales.get('Total Activo', 0)
    
    pasivo_corriente = sum([c['monto'] for c in report_data['Pasivo'].get('Pasivo Corriente', [])])
    pasivo_no_corriente = sum([c['monto'] for c in report_data['Pasivo'].get('Pasivo No Corriente', [])])
    total_pasivo = totales.get('Total Pasivo', 0)
    
    total_patrimonio = totales.get('Total Patrimonio', 0)
    
    ingresos = totales.get('Ingreso', 0)
    costos = totales.get('Costo', 0)
    gastos = totales.get('Gasto', 0)
    utilidad_bruta = totales.get('Utilidad Bruta', 0)
    utilidad_neta = totales.get('Utilidad Neta', 0)
    
    # Intentar obtener valores específicos de cuentas
    inventario = 0
    cuentas_por_cobrar = 0
    activos_fijos = 0
    
    for subtipo, cuentas in report_data['Activo'].items():
        for cuenta in cuentas:
            nombre_lower = cuenta['nombre'].lower()
            if 'inventario' in nombre_lower:
                inventario += cuenta['monto']
            if 'por cobrar' in nombre_lower or 'cuentas por cobrar' in nombre_lower:
                cuentas_por_cobrar += cuenta['monto']
            if 'propiedad' in nombre_lower or 'planta' in nombre_lower or 'equipo' in nombre_lower:
                activos_fijos += cuenta['monto']
    
    # === RATIOS DE LIQUIDEZ ===
    # Razón Circulante o Índice de Solvencia
    if pasivo_corriente > 0:
        razon_circulante = activo_corriente / pasivo_corriente
        estado_rc = 'optimo' if 1.5 <= razon_circulante <= 2.0 else ('alto' if razon_circulante > 2.0 else 'bajo')
        ratios['Liquidez']['Razón Circulante'] = {
            'nombre': 'Razón Circulante o Índice de Solvencia',
            'valor': razon_circulante,
            'formula': 'Activos Circulantes / Pasivo Circulante',
            'rango_optimo': '1.5 - 2.0',
            'estado': estado_rc,
            'interpretacion': f"Indica en qué medida los pasivos circulantes están cubiertos por los activos que se espera que se conviertan en efectivo en el futuro cercano. Un valor de {razon_circulante:.2f} {'sugiere que la empresa tiene suficiente capacidad para cubrir sus pasivos inmediatos' if razon_circulante > 1.0 else 'indica posibles problemas de liquidez'}."
        }
    
    # Razón Rápida
    if pasivo_corriente > 0:
        razon_rapida = (activo_corriente - inventario) / pasivo_corriente
        estado_rr = 'optimo' if razon_rapida >= 1.0 else 'bajo'
        ratios['Liquidez']['Razón Rápida'] = {
            'nombre': 'Razón Rápida',
            'valor': razon_rapida,
            'formula': '(Activos Circulantes - Inventarios) / Pasivo Circulante',
            'rango_optimo': '1.0 (aceptable y satisfactorio)',
            'estado': estado_rr,
            'interpretacion': f"La razón rápida mide la capacidad de la empresa para cubrir sus pasivos circulantes con sus activos más líquidos, excluyendo los inventarios. Un valor de {razon_rapida:.2f} {'es satisfactorio' if razon_rapida >= 1.0 else 'indica posibles problemas de liquidez inmediata'}.",
            'inventarios': inventario
        }
    
    # Capital de Trabajo
    capital_trabajo = activo_corriente - pasivo_corriente
    estado_ct = 'optimo' if capital_trabajo > 0 else 'bajo'
    ratios['Liquidez']['Capital de Trabajo'] = {
        'nombre': 'Capital de Trabajo',
        'valor': capital_trabajo,
        'formula': 'Activo Corriente - Pasivo Corriente',
        'rango_optimo': '> 0',
        'estado': estado_ct,
        'unidad': 'C$',
        'interpretacion': f"La empresa dispone de C${capital_trabajo:,.2f} en capital de trabajo."
    }
    
    # === RATIOS DE ACTIVIDAD ===
    # Rotación de Inventarios
    if inventario > 0 and costos > 0:
        rotacion_inventarios = costos / inventario
        estado_ri = 'optimo' if 5 <= rotacion_inventarios <= 10 else ('alto' if rotacion_inventarios > 10 else 'bajo')
        ratios['Actividades']['Rotación de Inventarios'] = {
            'nombre': 'Rotación de Inventarios',
            'valor': rotacion_inventarios,
            'formula': 'Costo de Bienes Vendidos / Inventarios Promedio',
            'rango_optimo': '5 - 10',
            'estado': estado_ri,
            'interpretacion': f"La rotación de inventarios muestra la eficiencia de la empresa en la venta y reposición de inventarios. Un valor de {rotacion_inventarios:.2f} {'indica que los inventarios se están utilizando eficientemente' if 5 <= rotacion_inventarios <= 10 else 'indica que los inventarios se están utilizando muy rápidamente' if rotacion_inventarios > 10 else 'indica que los inventarios se están utilizando lentamente'}."
        }
    
    # Rotación de Cuentas por Cobrar
    if cuentas_por_cobrar > 0 and ingresos > 0:
        rotacion_cuentas_cobrar = ingresos / cuentas_por_cobrar
        periodo_cobro = 360 / rotacion_cuentas_cobrar
        estado_rcc = 'optimo' if 6 <= rotacion_cuentas_cobrar <= 12 else ('alto' if rotacion_cuentas_cobrar > 12 else 'bajo')
        ratios['Actividades']['Rotación de Cuentas por Cobrar'] = {
            'nombre': 'Rotación de Cuentas por Cobrar',
            'valor': rotacion_cuentas_cobrar,
            'formula': 'Ventas al crédito / Cuentas por Cobrar Promedio',
            'rango_optimo': '6 - 12',
            'estado': estado_rcc,
            'interpretacion': f"La rotación de cuentas por cobrar mide la eficacia de la empresa en la gestión de cobros. Un valor de {rotacion_cuentas_cobrar:.2f} {'indica que la empresa cobra eficientemente a sus clientes' if 6 <= rotacion_cuentas_cobrar <= 12 else 'indica que la empresa cobra muy rápidamente' if rotacion_cuentas_cobrar > 12 else 'indica que la empresa cobra lentamente'}."
        }
        ratios['Actividades']['Periodo Promedio de Cobro'] = {
            'nombre': 'Periodo Promedio de Cobro',
            'valor': periodo_cobro,
            'formula': '360 / Rotación de Cuentas por Cobrar',
            'rango_optimo': '30 - 45 días',
            'estado': 'optimo' if 30 <= periodo_cobro <= 45 else ('alto' if periodo_cobro > 45 else 'bajo'),
            'unidad': 'días',
            'interpretacion': f"El periodo promedio de cobro indica el tiempo promedio que tarda la empresa en cobrar sus cuentas por cobrar. Un valor de {periodo_cobro:.0f} días {'es favorable' if periodo_cobro <= 45 else 'indica que se tarda más de lo óptimo en cobrar'}."
        }
    
    # Rotación de Activos Fijos
    if activos_fijos > 0 and ingresos > 0:
        rotacion_activos_fijos = ingresos / activos_fijos
        estado_raf = 'optimo' if 5 <= rotacion_activos_fijos <= 8 else ('alto' if rotacion_activos_fijos > 8 else 'bajo')
        ratios['Actividades']['Rotación de Activos Fijos'] = {
            'nombre': 'Rotación de Activos Fijos',
            'valor': rotacion_activos_fijos,
            'formula': 'Ventas / Activos Fijos Promedio',
            'rango_optimo': '5 - 8',
            'estado': estado_raf,
            'interpretacion': f"La rotación de activos fijos muestra cuán eficientemente la empresa utiliza sus activos fijos para generar ventas. Un valor de {rotacion_activos_fijos:.2f} {'indica mayor eficiencia' if rotacion_activos_fijos >= 5 else 'indica menor eficiencia'}."
        }
    
    # Rotación de Activos Totales
    if total_activo > 0 and ingresos > 0:
        rotacion_activos_totales = ingresos / total_activo
        estado_rat = 'optimo' if 1.0 <= rotacion_activos_totales <= 2.5 else ('alto' if rotacion_activos_totales > 2.5 else 'bajo')
        ratios['Actividades']['Rotación de Activos Totales'] = {
            'nombre': 'Rotación de Activos Totales',
            'valor': rotacion_activos_totales,
            'formula': 'Ventas / Activos Totales Promedio',
            'rango_optimo': '1.0 - 2.5',
            'estado': estado_rat,
            'interpretacion': f"La rotación de activos totales mide la eficacia con la que una empresa utiliza todos sus activos para generar ventas. Un valor de {rotacion_activos_totales:.2f} {'indica mayor eficiencia' if rotacion_activos_totales >= 1.0 else 'indica menor eficiencia'}."
        }
    
    # === RATIOS DE ENDEUDAMIENTO ===
    # Razón de Endeudamiento o deuda
    if total_activo > 0:
        razon_endeudamiento = total_pasivo / total_activo
        porcentaje_endeudamiento = razon_endeudamiento * 100
        estado_endeud = 'optimo' if 0.3 <= razon_endeudamiento <= 0.5 else ('alto' if razon_endeudamiento > 0.5 else 'bajo')
        ratios['Endeudamiento']['Razón de Endeudamiento'] = {
            'nombre': 'Razón de Endeudamiento o deuda',
            'valor': razon_endeudamiento,
            'formula': 'Total Pasivos / Total Activos',
            'rango_optimo': '0.3 - 0.5',
            'estado': estado_endeud,
            'porcentaje': porcentaje_endeudamiento,
            'interpretacion': f"La razón de endeudamiento indica el porcentaje de los activos que está financiado con deuda. Un valor de {porcentaje_endeudamiento:.1f}% {'es favorable, ya que indica menos dependencia de la deuda' if razon_endeudamiento <= 0.5 else 'implica un mayor riesgo de insolvencia'}."
        }
    
    # Razón Pasivo / Capital
    if total_patrimonio > 0:
        razon_patrimonio = total_pasivo / total_patrimonio
        estado_pat = 'optimo' if 0.5 <= razon_patrimonio <= 1.0 else ('alto' if razon_patrimonio > 1.0 else 'bajo')
        ratios['Endeudamiento']['Razón Pasivo / Capital'] = {
            'nombre': 'Razón Pasivo / Capital',
            'valor': razon_patrimonio,
            'formula': 'Total Pasivos / Patrimonio Neto',
            'rango_optimo': '0.5 - 1.0',
            'estado': estado_pat,
            'interpretacion': f"Esta razón indica la proporción de los activos financiados por deuda frente al capital propio. Un valor de {razon_patrimonio:.2f} {'es preferido' if razon_patrimonio <= 1.0 else 'indica mayor dependencia de deuda'}."
        }
    
    # Rotación de Intereses a Utilidades
    gastos_financieros = 0
    for subtipo, cuentas in report_data['Gasto'].items():
        for cuenta in cuentas:
            if 'financiero' in cuenta['nombre'].lower() or 'interes' in cuenta['nombre'].lower():
                gastos_financieros += abs(cuenta['monto'])  # Usar valor absoluto para gastos
    
    utilidad_operativa = totales.get('Utilidad Operativa', utilidad_bruta)
    if gastos_financieros > 0 and utilidad_operativa > 0:
        razon_intereses_utilidades = utilidad_operativa / gastos_financieros
        estado_riu = 'optimo' if 3 <= razon_intereses_utilidades <= 5 else ('alto' if razon_intereses_utilidades > 5 else 'bajo')
        ratios['Endeudamiento']['Rotación de Intereses a Utilidades'] = {
            'nombre': 'Rotación de Intereses a Utilidades',
            'valor': razon_intereses_utilidades,
            'formula': 'Utilidad Operativa / Gasto por Intereses',
            'rango_optimo': '3 - 5',
            'estado': estado_riu,
            'interpretacion': f"Esta razón mide la capacidad de la empresa para cubrir sus gastos por intereses con su utilidad operativa. Un valor de {razon_intereses_utilidades:.2f} {'indica mayor capacidad de pago de intereses' if razon_intereses_utilidades >= 3 else 'indica menor capacidad de pago'}."
        }
    
    # === RATIOS DE RENTABILIDAD ===
    # Margen de Utilidad Bruta (MUB)
    if ingresos > 0:
        margen_bruto = (utilidad_bruta / ingresos) * 100
        estado_mb = 'optimo' if 20 <= margen_bruto <= 40 else ('alto' if margen_bruto > 40 else 'bajo')
        ratios['Rentabilidad']['Margen de Utilidad Bruta (MUB)'] = {
            'nombre': 'Margen de Utilidad Bruta (MUB)',
            'valor': margen_bruto,
            'formula': 'Utilidad Bruta / Ventas',
            'rango_optimo': '20% - 40%',
            'estado': estado_mb,
            'interpretacion': f"El margen de utilidad bruta muestra la rentabilidad de la empresa antes de los gastos operativos. Un valor de {margen_bruto:.1f}% {'indica una mayor rentabilidad en la producción' if margen_bruto >= 20 else 'indica menor rentabilidad'}."
        }
    
    # Margen de Utilidad Operativa (MUO)
    utilidad_operativa = totales.get('Utilidad Operativa', utilidad_bruta)
    if ingresos > 0:
        margen_operativo = (utilidad_operativa / ingresos) * 100
        estado_mo = 'optimo' if 10 <= margen_operativo <= 20 else ('alto' if margen_operativo > 20 else 'bajo')
        ratios['Rentabilidad']['Margen de Utilidad Operativa (MUO)'] = {
            'nombre': 'Margen de Utilidad Operativa (MUO)',
            'valor': margen_operativo,
            'formula': 'Utilidad Operativa / Ventas',
            'rango_optimo': '10% - 20%',
            'estado': estado_mo,
            'interpretacion': f"El margen de utilidad operativa mide la rentabilidad de la empresa antes de los gastos financieros e impuestos. Un margen de {margen_operativo:.1f}% {'es favorable' if margen_operativo >= 10 else 'es bajo'}."
        }
    
    # Margen de Utilidad Neta
    if ingresos > 0:
        margen_neto = (utilidad_neta / ingresos) * 100
        estado_mn = 'optimo' if 5 <= margen_neto <= 10 else ('alto' if margen_neto > 10 else 'bajo')
        ratios['Rentabilidad']['Margen de Utilidad Neta'] = {
            'nombre': 'Margen de Utilidad Neta',
            'valor': margen_neto,
            'formula': 'Utilidad Neta / Ventas',
            'rango_optimo': '5% - 10%',
            'estado': estado_mn,
            'interpretacion': f"El margen de utilidad neta mide la rentabilidad final de la empresa, después de todos los gastos, impuestos e intereses. Un margen de {margen_neto:.1f}% {'indica una mayor rentabilidad para los accionistas' if margen_neto >= 5 else 'indica menor rentabilidad'}."
        }
    
    # Rentabilidad sobre el Activo (ROA)
    if total_activo > 0:
        roa = (utilidad_neta / total_activo) * 100
        estado_roa = 'optimo' if 5 <= roa <= 10 else ('alto' if roa > 10 else 'bajo')
        ratios['Rentabilidad']['Rentabilidad sobre el Activo (ROA)'] = {
            'nombre': 'Rentabilidad sobre el Activo (ROA)',
            'valor': roa,
            'formula': 'Utilidad Neta / Total Activos',
            'rango_optimo': '5% - 10%',
            'estado': estado_roa,
            'interpretacion': f"La rentabilidad sobre el activo mide la eficacia de la empresa para generar utilidades a partir de sus activos totales. Un ROA de {roa:.1f}% {'indica una mayor eficiencia en el uso de los activos' if roa >= 5 else 'indica menor eficiencia'}."
        }
    
    # ROE (Return on Equity)
    if total_patrimonio > 0:
        roe = (utilidad_neta / total_patrimonio) * 100
        estado_roe = 'optimo' if roe >= 15 else ('normal' if roe >= 10 else 'bajo')
        ratios['Rentabilidad']['ROE'] = {
            'nombre': 'ROE (Retorno sobre Patrimonio)',
            'valor': roe,
            'formula': '(Utilidad Neta / Patrimonio) × 100',
            'rango_optimo': '> 15%',
            'estado': estado_roe,
            'interpretacion': f"El patrimonio genera un {roe:.1f}% de retorno para los accionistas."
        }
    
    return ratios

def calcular_origen_aplicacion(report_data_base, report_data_analisis):
    """
    Calcula el origen y aplicación de fondos comparando dos períodos.
    Origen: Aumentos en Pasivo y Patrimonio, Disminuciones en Activo
    Aplicación: Aumentos en Activo, Disminuciones en Pasivo y Patrimonio
    """
    origen_aplicacion = {
        'Origen': defaultdict(list),
        'Aplicacion': defaultdict(list),
        'Totales': {
            'Origen': {'Total': 0.0},
            'Aplicacion': {'Total': 0.0}
        }
    }
    
    # Obtener nombres de cuentas desde el catálogo para casos donde no estén en los datos
    nombres_cuentas = {}
    try:
        with engine.connect() as conn:
            query_nombres = text("SELECT CuentaID, NombreCuenta FROM CatalogoCuentas")
            resultados_nombres = conn.execute(query_nombres).fetchall()
            nombres_cuentas = {row[0]: row[1] for row in resultados_nombres}
    except Exception as e:
        print(f"Error al obtener nombres de cuentas: {e}")
    
    # Procesar Activos (Aumento = Aplicación, Disminución = Origen)
    for subtipo in set(list(report_data_base['Activo'].keys()) + list(report_data_analisis['Activo'].keys())):
        cuentas_base = {c['id']: c for c in report_data_base['Activo'].get(subtipo, [])}
        cuentas_analisis = {c['id']: c for c in report_data_analisis['Activo'].get(subtipo, [])}
        
        todas_cuentas_ids = set(cuentas_base.keys()) | set(cuentas_analisis.keys())
        
        for cuenta_id in todas_cuentas_ids:
            cuenta_base_obj = cuentas_base.get(cuenta_id, {'nombre': '', 'monto': 0.0})
            cuenta_analisis_obj = cuentas_analisis.get(cuenta_id, {'nombre': cuenta_base_obj.get('nombre', ''), 'monto': 0.0})
            
            monto_base = float(cuenta_base_obj.get('monto', 0.0))
            monto_analisis = float(cuenta_analisis_obj.get('monto', 0.0))
            variacion = monto_analisis - monto_base
            
            # Obtener el nombre de la cuenta, priorizando el del análisis, luego el base, luego el catálogo
            nombre_cuenta = (cuenta_analisis_obj.get('nombre', '') or 
                           cuenta_base_obj.get('nombre', '') or 
                           nombres_cuentas.get(cuenta_id, '') or 
                           f'Cuenta {cuenta_id}')
            
            # Solo agregar si hay variación (diferente de cero) o si hay datos en al menos un período
            if abs(variacion) > 0.01 or monto_base != 0.0 or monto_analisis != 0.0:
                if variacion > 0:
                    # Aumento en Activo = Aplicación
                    origen_aplicacion['Aplicacion'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Activo',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(variacion)
                    })
                    origen_aplicacion['Totales']['Aplicacion']['Total'] += variacion
                elif variacion < 0:
                    # Disminución en Activo = Origen
                    origen_aplicacion['Origen'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Activo',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(abs(variacion))
                    })
                    origen_aplicacion['Totales']['Origen']['Total'] += abs(variacion)
    
    # Procesar Pasivos (Aumento = Origen, Disminución = Aplicación)
    for subtipo in set(list(report_data_base['Pasivo'].keys()) + list(report_data_analisis['Pasivo'].keys())):
        cuentas_base = {c['id']: c for c in report_data_base['Pasivo'].get(subtipo, [])}
        cuentas_analisis = {c['id']: c for c in report_data_analisis['Pasivo'].get(subtipo, [])}
        
        todas_cuentas_ids = set(cuentas_base.keys()) | set(cuentas_analisis.keys())
        
        for cuenta_id in todas_cuentas_ids:
            cuenta_base_obj = cuentas_base.get(cuenta_id, {'nombre': '', 'monto': 0.0})
            cuenta_analisis_obj = cuentas_analisis.get(cuenta_id, {'nombre': cuenta_base_obj.get('nombre', ''), 'monto': 0.0})
            
            monto_base = float(cuenta_base_obj.get('monto', 0.0))
            monto_analisis = float(cuenta_analisis_obj.get('monto', 0.0))
            variacion = monto_analisis - monto_base
            
            # Obtener el nombre de la cuenta, priorizando el del análisis, luego el base, luego el catálogo
            nombre_cuenta = (cuenta_analisis_obj.get('nombre', '') or 
                           cuenta_base_obj.get('nombre', '') or 
                           nombres_cuentas.get(cuenta_id, '') or 
                           f'Cuenta {cuenta_id}')
            
            # Solo agregar si hay variación (diferente de cero) o si hay datos en al menos un período
            if abs(variacion) > 0.01 or monto_base != 0.0 or monto_analisis != 0.0:
                if variacion > 0:
                    # Aumento en Pasivo = Origen
                    origen_aplicacion['Origen'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Pasivo',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(variacion)
                    })
                    origen_aplicacion['Totales']['Origen']['Total'] += variacion
                elif variacion < 0:
                    # Disminución en Pasivo = Aplicación
                    origen_aplicacion['Aplicacion'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Pasivo',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(abs(variacion))
                    })
                    origen_aplicacion['Totales']['Aplicacion']['Total'] += abs(variacion)
    
    # Procesar Patrimonio (Aumento = Origen, Disminución = Aplicación)
    for subtipo in set(list(report_data_base['Patrimonio'].keys()) + list(report_data_analisis['Patrimonio'].keys())):
        cuentas_base = {c['id']: c for c in report_data_base['Patrimonio'].get(subtipo, [])}
        cuentas_analisis = {c['id']: c for c in report_data_analisis['Patrimonio'].get(subtipo, [])}
        
        todas_cuentas_ids = set(cuentas_base.keys()) | set(cuentas_analisis.keys())
        
        for cuenta_id in todas_cuentas_ids:
            cuenta_base_obj = cuentas_base.get(cuenta_id, {'nombre': '', 'monto': 0.0})
            cuenta_analisis_obj = cuentas_analisis.get(cuenta_id, {'nombre': cuenta_base_obj.get('nombre', ''), 'monto': 0.0})
            
            monto_base = float(cuenta_base_obj.get('monto', 0.0))
            monto_analisis = float(cuenta_analisis_obj.get('monto', 0.0))
            variacion = monto_analisis - monto_base
            
            # Obtener el nombre de la cuenta, priorizando el del análisis, luego el base, luego el catálogo
            nombre_cuenta = (cuenta_analisis_obj.get('nombre', '') or 
                           cuenta_base_obj.get('nombre', '') or 
                           nombres_cuentas.get(cuenta_id, '') or 
                           f'Cuenta {cuenta_id}')
            
            # Solo agregar si hay variación (diferente de cero) o si hay datos en al menos un período
            if abs(variacion) > 0.01 or monto_base != 0.0 or monto_analisis != 0.0:
                if variacion > 0:
                    # Aumento en Patrimonio = Origen
                    origen_aplicacion['Origen'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Patrimonio',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(variacion)
                    })
                    origen_aplicacion['Totales']['Origen']['Total'] += variacion
                elif variacion < 0:
                    # Disminución en Patrimonio = Aplicación
                    origen_aplicacion['Aplicacion'][subtipo].append({
                        'id': cuenta_id,
                        'nombre': nombre_cuenta,
                        'tipo': 'Patrimonio',
                        'monto_base': float(monto_base),
                        'monto_analisis': float(monto_analisis),
                        'variacion': float(abs(variacion))
                    })
                    origen_aplicacion['Totales']['Aplicacion']['Total'] += abs(variacion)
    
    # Calcular totales por subtipo para Origen y Aplicación
    for subtipo in origen_aplicacion['Origen'].keys():
        total_subtipo = sum(cuenta.get('variacion', 0.0) for cuenta in origen_aplicacion['Origen'][subtipo])
        origen_aplicacion['Totales']['Origen'][subtipo] = total_subtipo
    
    for subtipo in origen_aplicacion['Aplicacion'].keys():
        total_subtipo = sum(cuenta.get('variacion', 0.0) for cuenta in origen_aplicacion['Aplicacion'][subtipo])
        origen_aplicacion['Totales']['Aplicacion'][subtipo] = total_subtipo
    
    return origen_aplicacion

# --- Funciones para análisis con IA (Gemini) ---

def analizar_con_gemini(report_data, anio_seleccionado):
    """Genera un análisis financiero pequeño y enfocado usando Gemini"""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        
        # Extraer datos clave
        totales = report_data.get('Totales', {})
        total_activo = totales.get('Total Activo', 0)
        total_pasivo = totales.get('Total Pasivo', 0)
        total_patrimonio = totales.get('Total Patrimonio', 0)
        ingresos = totales.get('Ingreso', 0)
        costos = totales.get('Costo', 0)
        gastos = totales.get('Gasto', 0)
        utilidad_bruta = totales.get('Utilidad Bruta', 0)
        utilidad_operativa = totales.get('Utilidad Operativa', 0)
        utilidad_neta = totales.get('Utilidad Neta', 0)
        
        # Calcular ratios básicos para contexto
        razon_pasivo_patrimonio = (total_pasivo / total_patrimonio) if total_patrimonio > 0 else 0
        margen_neto = (utilidad_neta / ingresos * 100) if ingresos > 0 else 0
        
        # Prompt conciso y estructurado
        prompt = f"""Como analista financiero, proporciona un resumen ejecutivo breve (máximo 4 puntos clave) del análisis vertical para el año {anio_seleccionado}.

Datos principales del Balance General:
- Total Activo: C${total_activo:,.2f}
- Total Pasivo: C${total_pasivo:,.2f}
- Total Patrimonio: C${total_patrimonio:,.2f}
- Razón Pasivo/Patrimonio: {razon_pasivo_patrimonio:.2f}

Datos principales del Estado de Resultados:
- Ingresos: C${ingresos:,.2f}
- Costos: C${costos:,.2f}
- Gastos: C${gastos:,.2f}
- Utilidad Bruta: C${utilidad_bruta:,.2f}
- Utilidad Operativa: C${utilidad_operativa:,.2f}
- Utilidad Neta: C${utilidad_neta:,.2f}
- Margen Neto: {margen_neto:.1f}%

Enfoca tu análisis en:
1. Estructura de capital (relación pasivo/patrimonio y su significado)
2. Rentabilidad (margen neto y eficiencia operativa)
3. Principales fortalezas identificadas en la estructura financiera
4. Recomendación clave basada en los datos presentados

Sé directo, específico y usa las cifras proporcionadas. Máximo 80 palabras."""

        response = model.generate_content(prompt)
        analisis_texto = response.text
        
        # Convertir markdown a HTML
        analisis_html = markdown(analisis_texto)
        
        return analisis_html
        
    except Exception as e:
        print(f"Error al llamar a la API de Gemini: {e}")
        import traceback
        traceback.print_exc()
        return f"<p><strong>Error al generar el análisis:</strong> {str(e)}</p>"

def analizar_horizontal_ia(report_data_base, report_data_analisis, periodo_base, periodo_analisis):
    """Análisis horizontal pequeño y enfocado"""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        
        # Obtener totales de ambos períodos
        totales_base = report_data_base.get('Totales', {})
        totales_analisis = report_data_analisis.get('Totales', {})
        
        # Calcular variaciones principales
        activo_base = totales_base.get('Total Activo', 0)
        activo_analisis = totales_analisis.get('Total Activo', 0)
        activo_var = activo_analisis - activo_base
        activo_var_pct = (activo_var / activo_base * 100) if activo_base > 0 else 0
        
        pasivo_base = totales_base.get('Total Pasivo', 0)
        pasivo_analisis = totales_analisis.get('Total Pasivo', 0)
        pasivo_var = pasivo_analisis - pasivo_base
        pasivo_var_pct = (pasivo_var / pasivo_base * 100) if pasivo_base > 0 else 0
        
        patrimonio_base = totales_base.get('Total Patrimonio', 0)
        patrimonio_analisis = totales_analisis.get('Total Patrimonio', 0)
        patrimonio_var = patrimonio_analisis - patrimonio_base
        patrimonio_var_pct = (patrimonio_var / patrimonio_base * 100) if patrimonio_base > 0 else 0
        
        ingresos_base = totales_base.get('Ingreso', 0)
        ingresos_analisis = totales_analisis.get('Ingreso', 0)
        ingresos_var = ingresos_analisis - ingresos_base
        ingresos_var_pct = (ingresos_var / ingresos_base * 100) if ingresos_base > 0 else 0
        
        utilidad_base = totales_base.get('Utilidad Neta', 0)
        utilidad_analisis = totales_analisis.get('Utilidad Neta', 0)
        utilidad_var = utilidad_analisis - utilidad_base
        utilidad_var_pct = (utilidad_var / utilidad_base * 100) if utilidad_base != 0 else (100 if utilidad_var > 0 else -100 if utilidad_var < 0 else 0)
        
        prompt = f"""Como analista financiero, proporciona un análisis horizontal breve (máximo 4 puntos clave) comparando el período {periodo_base} vs {periodo_analisis}.

Variaciones en Balance General:
- Activo Total: {activo_var_pct:+.1f}% (C${activo_var:+,.2f}) - De C${activo_base:,.2f} a C${activo_analisis:,.2f}
- Pasivo Total: {pasivo_var_pct:+.1f}% (C${pasivo_var:+,.2f}) - De C${pasivo_base:,.2f} a C${pasivo_analisis:,.2f}
- Patrimonio: {patrimonio_var_pct:+.1f}% (C${patrimonio_var:+,.2f}) - De C${patrimonio_base:,.2f} a C${patrimonio_analisis:,.2f}

Variaciones en Estado de Resultados:
- Ingresos: {ingresos_var_pct:+.1f}% (C${ingresos_var:+,.2f}) - De C${ingresos_base:,.2f} a C${ingresos_analisis:,.2f}
- Utilidad Neta: {utilidad_var_pct:+.1f}% (C${utilidad_var:+,.2f}) - De C${utilidad_base:,.2f} a C${utilidad_analisis:,.2f}

Enfoca tu análisis en:
1. Cambio más significativo y su impacto
2. Evolución de la estructura financiera (activos, pasivos, patrimonio)
3. Tendencia de rentabilidad (mejora o deterioro)
4. Conclusión práctica y recomendación clave

Sé directo, específico y usa las cifras proporcionadas. Máximo 80 palabras."""

        response = model.generate_content(prompt)
        analisis_texto = response.text
        
        # Convertir markdown a HTML
        analisis_html = markdown(analisis_texto)
        
        return analisis_html
    except Exception as e:
        print(f"Error en análisis horizontal IA: {e}")
        import traceback
        traceback.print_exc()
        return f"<p><strong>Error al generar el análisis:</strong> {str(e)}</p>"

def analizar_ratios_ia(ratios_data):
    """Análisis de ratios pequeño y enfocado"""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        
        # Extraer ratios clave de la estructura correcta
        razon_circulante = ratios_data.get('Liquidez', {}).get('Razón Circulante', {}).get('valor', 0)
        razon_rapida = ratios_data.get('Liquidez', {}).get('Razón Rápida', {}).get('valor', 0)
        margen_bruto = ratios_data.get('Rentabilidad', {}).get('Margen de Utilidad Bruta (MUB)', {}).get('valor', 0)
        margen_neto = ratios_data.get('Rentabilidad', {}).get('Margen de Utilidad Neta', {}).get('valor', 0)
        roa = ratios_data.get('Rentabilidad', {}).get('Rentabilidad sobre el Activo (ROA)', {}).get('valor', 0)
        razon_endeudamiento = ratios_data.get('Endeudamiento', {}).get('Razón de Endeudamiento', {}).get('valor', 0)
        razon_endeudamiento_pct = ratios_data.get('Endeudamiento', {}).get('Razón de Endeudamiento', {}).get('porcentaje', 0)
        
        # Obtener estados para contexto
        estado_liquidez = ratios_data.get('Liquidez', {}).get('Razón Circulante', {}).get('estado', 'normal')
        estado_rentabilidad = ratios_data.get('Rentabilidad', {}).get('Margen de Utilidad Neta', {}).get('estado', 'normal')
        estado_endeudamiento = ratios_data.get('Endeudamiento', {}).get('Razón de Endeudamiento', {}).get('estado', 'normal')
        
        prompt = f"""Como analista financiero, proporciona un análisis ejecutivo breve (máximo 4 puntos clave) de los ratios financieros calculados.

Datos principales de ratios:
- Razón Circulante: {razon_circulante:.2f} (Estado: {estado_liquidez})
- Razón Rápida: {razon_rapida:.2f}
- Margen Bruto: {margen_bruto:.1f}%
- Margen Neto: {margen_neto:.1f}% (Estado: {estado_rentabilidad})
- ROA: {roa:.1f}%
- Razón de Endeudamiento: {razon_endeudamiento:.2f} ({razon_endeudamiento_pct:.1f}%) (Estado: {estado_endeudamiento})

Enfoca tu análisis en:
1. Situación de liquidez (capacidad de pago a corto plazo)
2. Rentabilidad (eficiencia en generación de utilidades)
3. Endeudamiento (estructura de capital y riesgo)
4. Recomendación clave basada en los ratios calculados

Sé directo, específico y usa las cifras proporcionadas. Máximo 80 palabras."""

        response = model.generate_content(prompt)
        analisis_texto = response.text
        
        # Convertir markdown a HTML
        analisis_html = markdown(analisis_texto)
        
        return analisis_html
    except Exception as e:
        print(f"Error en análisis ratios IA: {e}")
        import traceback
        traceback.print_exc()
        return f"<p><strong>Error al generar el análisis:</strong> {str(e)}</p>"

def analizar_origen_aplicacion_ia(origen_aplicacion_data):
    """Análisis de origen y aplicación pequeño y enfocado"""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        
        # Obtener totales de forma segura
        totales = origen_aplicacion_data.get('Totales', {})
        total_origen = totales.get('Origen', {}).get('Total', 0)
        total_aplicacion = totales.get('Aplicacion', {}).get('Total', 0)
        diferencia = total_origen - total_aplicacion
        
        # Identificar principales orígenes y aplicaciones
        principales_origenes = []
        origen_dict = origen_aplicacion_data.get('Origen', {})
        for subtipo, cuentas in origen_dict.items():
            if isinstance(cuentas, list):
                for cuenta in sorted(cuentas, key=lambda x: abs(x.get('variacion', 0)), reverse=True)[:3]:
                    nombre = cuenta.get('nombre', 'Sin nombre')
                    variacion = cuenta.get('variacion', 0)
                    principales_origenes.append(f"- {nombre}: C${variacion:,.2f}")
        
        principales_aplicaciones = []
        aplicacion_dict = origen_aplicacion_data.get('Aplicacion', {})
        for subtipo, cuentas in aplicacion_dict.items():
            if isinstance(cuentas, list):
                for cuenta in sorted(cuentas, key=lambda x: abs(x.get('variacion', 0)), reverse=True)[:3]:
                    nombre = cuenta.get('nombre', 'Sin nombre')
                    variacion = cuenta.get('variacion', 0)
                    principales_aplicaciones.append(f"- {nombre}: C${variacion:,.2f}")
        
        # Limitar a los 3 principales
        principales_origenes = principales_origenes[:3]
        principales_aplicaciones = principales_aplicaciones[:3]
        
        prompt = f"""Como analista financiero, proporciona un análisis breve (máximo 4 puntos clave) del Origen y Aplicación de Fondos.

Resumen de flujo de fondos:
- Total Origen de Fondos: C${total_origen:,.2f}
- Total Aplicación de Fondos: C${total_aplicacion:,.2f}
- Diferencia (Origen - Aplicación): C${diferencia:,.2f}

Principales orígenes de fondos:
{chr(10).join(principales_origenes) if principales_origenes else '- No hay orígenes significativos'}

Principales aplicaciones de fondos:
{chr(10).join(principales_aplicaciones) if principales_aplicaciones else '- No hay aplicaciones significativas'}

Enfoca tu análisis en:
1. Principales fuentes de recursos (orígenes más importantes)
2. Principales usos de recursos (aplicaciones más importantes)
3. Evaluación del equilibrio entre origen y aplicación
4. Conclusión práctica sobre la gestión del flujo de fondos

Sé directo, específico y usa las cifras proporcionadas. Máximo 80 palabras."""

        response = model.generate_content(prompt)
        analisis_texto = response.text
        
        # Convertir markdown a HTML
        analisis_html = markdown(analisis_texto)
        
        return analisis_html
    except Exception as e:
        print(f"Error en análisis origen/aplicación IA: {e}")
        import traceback
        traceback.print_exc()
        return f"<p><strong>Error al generar el análisis:</strong> {str(e)}</p>"

def analizar_flujo_efectivo_ia(flujo_data, periodo_inicio, periodo_fin):
    """Análisis de flujo de efectivo pequeño y enfocado"""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel('gemini-2.0-flash-lite')
        
        # Extraer totales
        fne_operacion = flujo_data.get('Operacion', {}).get('total', 0)
        fne_inversion = flujo_data.get('Inversion', {}).get('total', 0)
        fne_financiamiento = flujo_data.get('Financiamiento', {}).get('total', 0)
        flujo_neto_total = flujo_data.get('Validacion', {}).get('flujo_neto', 0)
        
        # Identificar principales movimientos (top 2 de cada sección)
        top_operacion = []
        detalles_op = flujo_data.get('Operacion', {}).get('detalles', [])
        if detalles_op:
            for item in sorted(detalles_op, key=lambda x: abs(x.get('monto', 0)), reverse=True)[:2]:
                top_operacion.append(f"- {item.get('concepto')}: C${item.get('monto', 0):,.2f}")
                
        top_inversion = []
        detalles_inv = flujo_data.get('Inversion', {}).get('detalles', [])
        if detalles_inv:
            for item in sorted(detalles_inv, key=lambda x: abs(x.get('monto', 0)), reverse=True)[:2]:
                top_inversion.append(f"- {item.get('concepto')}: C${item.get('monto', 0):,.2f}")
                
        top_financiamiento = []
        detalles_fin = flujo_data.get('Financiamiento', {}).get('detalles', [])
        if detalles_fin:
            for item in sorted(detalles_fin, key=lambda x: abs(x.get('monto', 0)), reverse=True)[:2]:
                top_financiamiento.append(f"- {item.get('concepto')}: C${item.get('monto', 0):,.2f}")
        
        prompt = f"""Como analista financiero, proporciona un análisis ejecutivo breve (máximo 4 puntos clave) del Estado de Flujo de Efectivo (Método Indirecto) para el período {periodo_inicio}-{periodo_fin}.

Resumen de Flujos:
- Actividades de Operación: C${fne_operacion:,.2f}
  {chr(10).join(top_operacion)}
- Actividades de Inversión: C${fne_inversion:,.2f}
  {chr(10).join(top_inversion)}
- Actividades de Financiamiento: C${fne_financiamiento:,.2f}
  {chr(10).join(top_financiamiento)}
- Flujo Neto Total: C${flujo_neto_total:,.2f}

Enfoca tu análisis en:
1. Capacidad de generación de efectivo operativo (calidad de las utilidades)
2. Estrategia de inversión y crecimiento (o desinversión)
3. Política de financiamiento (dependencia de deuda vs capital)
4. Conclusión sobre la sostenibilidad de la liquidez

Sé directo, específico y usa las cifras proporcionadas. Máximo 80 palabras."""

        response = model.generate_content(prompt)
        analisis_texto = response.text
        
        # Convertir markdown a HTML
        analisis_html = markdown(analisis_texto)
        
        return analisis_html
    except Exception as e:
        print(f"Error en análisis flujo efectivo IA: {e}")
        import traceback
        traceback.print_exc()
        return f"<p><strong>Error al generar el análisis:</strong> {str(e)}</p>"

def exportar_analisis_vertical_excel(anio_seleccionado, report_data):
    """Exporta solo el Análisis Vertical a Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Acceso seguro a los datos
        totales = report_data.get('Totales', {})
        base_bg = totales.get('Total Activo', 0)
        base_er = totales.get('Ingreso', 0)
        
        ws = wb.create_sheet("Análisis Vertical")
        ws.append([f"ANÁLISIS VERTICAL - AÑO {anio_seleccionado}"])
        ws.append([])
        
        # Balance General
        ws.append(["BALANCE GENERAL (Base: Total Activos)"])
        ws.append(["Cuenta", "Monto (C$)", "% Vertical"])
        
        # Activo
        activo_data = report_data.get('Activo', {})
        for subtipo, cuentas in activo_data.items():
            if cuentas:  # Solo si hay cuentas
                ws.append([f"  {subtipo}"])
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_bg * 100) if base_bg > 0 else 0
                        ws.append([f"    {nombre}", monto, porcentaje])
        
        # Pasivo
        pasivo_data = report_data.get('Pasivo', {})
        for subtipo, cuentas in pasivo_data.items():
            if cuentas:  # Solo si hay cuentas
                ws.append([f"  {subtipo}"])
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_bg * 100) if base_bg > 0 else 0
                        ws.append([f"    {nombre}", monto, porcentaje])
        
        # Patrimonio
        patrimonio_data = report_data.get('Patrimonio', {})
        for subtipo, cuentas in patrimonio_data.items():
            if cuentas:  # Solo si hay cuentas
                ws.append([f"  {subtipo}"])
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_bg * 100) if base_bg > 0 else 0
                        ws.append([f"    {nombre}", monto, porcentaje])
        ws.append([])
        
        # Estado de Resultados
        ws.append(["ESTADO DE RESULTADOS (Base: Ingresos)"])
        ws.append(["Cuenta", "Monto (C$)", "% Vertical"])
        
        # Ingresos
        ingreso_data = report_data.get('Ingreso', {})
        for subtipo, cuentas in ingreso_data.items():
            if cuentas:
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_er * 100) if base_er > 0 else 0
                        ws.append([nombre, monto, porcentaje])
        
        # Costos
        costo_data = report_data.get('Costo', {})
        for subtipo, cuentas in costo_data.items():
            if cuentas:
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_er * 100) if base_er > 0 else 0
                        ws.append([nombre, monto, porcentaje])
        
        # Gastos
        gasto_data = report_data.get('Gasto', {})
        for subtipo, cuentas in gasto_data.items():
            if cuentas:
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        monto = cuenta.get('monto', 0)
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        porcentaje = (monto / base_er * 100) if base_er > 0 else 0
                        ws.append([nombre, monto, porcentaje])
        
        # Aplicar estilos
        ws['A1'].font = title_font
        ws.merge_cells('A1:C1')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row <= 3 or (cell.value and isinstance(cell.value, str) and ('BASE' in str(cell.value).upper() or cell.value.strip().startswith('  '))):
                    if cell.row <= 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = center_alignment
                elif cell.column in [2, 3] and isinstance(cell.value, (int, float)):
                    cell.alignment = right_alignment
                    if cell.column == 3:
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = '#,##0.00'
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        
        return wb
    except Exception as e:
        print(f"Error al exportar Análisis Vertical: {e}")
        return None

def exportar_analisis_horizontal_excel(periodo_base, periodo_analisis, analisis_comparativo):
    """Exporta solo el Análisis Horizontal a Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        ws = wb.create_sheet("Análisis Horizontal")
        ws.append([f"ANÁLISIS HORIZONTAL - {periodo_base} vs {periodo_analisis}"])
        ws.append([])
        
        # Encabezados
        ws.append(["Cuenta", f"Período Base ({periodo_base})", f"Período Análisis ({periodo_analisis})", "Variación Absoluta", "Variación Relativa (%)"])
        
        # Balance General
        for tipo in ['Activo', 'Pasivo', 'Patrimonio']:
            tipo_data = analisis_comparativo.get(tipo, {})
            if tipo_data:
                ws.append([f"  {tipo}"])
                for subtipo, cuentas in tipo_data.items():
                    if cuentas:
                        ws.append([f"    {subtipo}"])
                        for cuenta in cuentas:
                            if isinstance(cuenta, dict):
                                nombre = cuenta.get('nombre', 'Sin nombre')
                                monto_base = cuenta.get('monto_base', 0)
                                monto_analisis = cuenta.get('monto_analisis', 0)
                                absoluto = cuenta.get('absoluto', 0)
                                relativo = cuenta.get('relativo', 0)
                                # Manejar infinito
                                if relativo == float('inf'):
                                    relativo_str = '∞'
                                elif relativo == float('-inf'):
                                    relativo_str = '-∞'
                                else:
                                    relativo_str = relativo
                                ws.append([
                                    f"      {nombre}",
                                    monto_base,
                                    monto_analisis,
                                    absoluto,
                                    relativo_str
                                ])
        
        # Estado de Resultados
        for tipo in ['Ingreso', 'Costo', 'Gasto']:
            tipo_data = analisis_comparativo.get(tipo, {})
            if tipo_data:
                ws.append([f"  {tipo}"])
                for subtipo, cuentas in tipo_data.items():
                    if cuentas:
                        for cuenta in cuentas:
                            if isinstance(cuenta, dict):
                                nombre = cuenta.get('nombre', 'Sin nombre')
                                monto_base = cuenta.get('monto_base', 0)
                                monto_analisis = cuenta.get('monto_analisis', 0)
                                absoluto = cuenta.get('absoluto', 0)
                                relativo = cuenta.get('relativo', 0)
                                # Manejar infinito
                                if relativo == float('inf'):
                                    relativo_str = '∞'
                                elif relativo == float('-inf'):
                                    relativo_str = '-∞'
                                else:
                                    relativo_str = relativo
                                ws.append([
                                    f"    {nombre}",
                                    monto_base,
                                    monto_analisis,
                                    absoluto,
                                    relativo_str
                                ])
        
        # Totales
        ws.append([])
        ws.append(["TOTALES"])
        totales_data = analisis_comparativo.get('Totales', {})
        for key, valor in totales_data.items():
            if isinstance(valor, dict):
                base = valor.get('base', 0)
                analisis = valor.get('analisis', 0)
                absoluto = valor.get('absoluto', 0)
                relativo = valor.get('relativo', 0)
                # Manejar infinito
                if relativo == float('inf'):
                    relativo_str = '∞'
                elif relativo == float('-inf'):
                    relativo_str = '-∞'
                else:
                    relativo_str = relativo
                ws.append([
                    key,
                    base,
                    analisis,
                    absoluto,
                    relativo_str
                ])
        
        # Aplicar estilos
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row <= 3:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                elif cell.column in [2, 3, 4, 5] and isinstance(cell.value, (int, float)):
                    cell.alignment = right_alignment
                    if cell.column == 5:
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = '#,##0.00'
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
        
        return wb
    except Exception as e:
        print(f"Error al exportar Análisis Horizontal: {e}")
        return None

def exportar_ratios_excel(anio_seleccionado, ratios_data):
    """Exporta solo los Ratios Financieros a Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        ws = wb.create_sheet("Ratios Financieros")
        ws.append([f"RATIOS FINANCIEROS - AÑO {anio_seleccionado}"])
        ws.append([])
        
        # Encabezados
        ws.append(["Categoría", "Ratio", "Valor", "Fórmula", "Rango Óptimo", "Estado", "Interpretación"])
        
        # Acceder a la estructura correcta por categorías
        categorias = {
            'Liquidez': ratios_data.get('Liquidez', {}),
            'Actividades': ratios_data.get('Actividades', {}),
            'Endeudamiento': ratios_data.get('Endeudamiento', {}),
            'Rentabilidad': ratios_data.get('Rentabilidad', {})
        }
        
        for categoria_nombre, categoria_ratios in categorias.items():
            if categoria_ratios:
                for ratio_nombre, ratio_info in categoria_ratios.items():
                    if isinstance(ratio_info, dict):
                        valor = ratio_info.get('valor', 0)
                        formula = ratio_info.get('formula', '')
                        rango_optimo = ratio_info.get('rango_optimo', '')
                        estado = ratio_info.get('estado', 'normal')
                        interpretacion = ratio_info.get('interpretacion', '')
                        
                        # Formatear valor según el tipo
                        if 'unidad' in ratio_info:
                            if ratio_info['unidad'] == 'días':
                                valor_str = f"{valor:.0f} días"
                            elif ratio_info['unidad'] == 'C$':
                                valor_str = f"C${valor:,.2f}"
                            else:
                                valor_str = f"{valor:.2f} {ratio_info['unidad']}"
                        elif 'porcentaje' in ratio_info:
                            valor_str = f"{ratio_info['porcentaje']:.1f}%"
                        else:
                            # Determinar si es porcentaje o número
                            if 'Margen' in ratio_nombre or 'ROA' in ratio_nombre or 'ROE' in ratio_nombre:
                                valor_str = f"{valor:.2f}%"
                            else:
                                valor_str = f"{valor:.2f}"
                        
                        ws.append([
                            categoria_nombre,
                            ratio_nombre,
                            valor_str,
                            formula,
                            rango_optimo,
                            estado.capitalize(),
                            interpretacion
                        ])
        
        # Aplicar estilos
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row <= 3:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                elif cell.column == 3 and isinstance(cell.value, str):
                    cell.alignment = right_alignment
                elif cell.column in [4, 5, 6]:
                    cell.alignment = center_alignment
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 60
        
        return wb
    except Exception as e:
        print(f"Error al exportar Ratios Financieros: {e}")
        import traceback
        traceback.print_exc()
        return None

def exportar_origen_aplicacion_excel(periodo_base, periodo_analisis, origen_aplicacion_data):
    """Exporta solo el Origen y Aplicación de Fondos a Excel"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        ws = wb.create_sheet("Origen y Aplicación")
        ws.append([f"ORIGEN Y APLICACIÓN DE FONDOS - {periodo_base} vs {periodo_analisis}"])
        ws.append([])
        
        # Origen de Fondos
        ws.append(["ORIGEN DE FONDOS"])
        ws.append(["Cuenta", "Monto Base", "Monto Análisis", "Variación (C$)"])
        origen_data = origen_aplicacion_data.get('Origen', {})
        for subtipo, cuentas in origen_data.items():
            if cuentas:
                ws.append([f"  {subtipo}"])
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        monto_base = cuenta.get('monto_base', 0)
                        monto_analisis = cuenta.get('monto_analisis', 0)
                        variacion = cuenta.get('variacion', 0)
                        ws.append([f"    {nombre}", monto_base, monto_analisis, variacion])
        ws.append([])
        
        # Aplicación de Fondos
        ws.append(["APLICACIÓN DE FONDOS"])
        ws.append(["Cuenta", "Monto Base", "Monto Análisis", "Variación (C$)"])
        aplicacion_data = origen_aplicacion_data.get('Aplicacion', {})
        for subtipo, cuentas in aplicacion_data.items():
            if cuentas:
                ws.append([f"  {subtipo}"])
                for cuenta in cuentas:
                    if isinstance(cuenta, dict):
                        nombre = cuenta.get('nombre', 'Sin nombre')
                        monto_base = cuenta.get('monto_base', 0)
                        monto_analisis = cuenta.get('monto_analisis', 0)
                        variacion = cuenta.get('variacion', 0)
                        ws.append([f"    {nombre}", monto_base, monto_analisis, variacion])
        ws.append([])
        
        # Totales
        ws.append(["TOTALES"])
        totales = origen_aplicacion_data.get('Totales', {})
        total_origen = totales.get('Origen', {}).get('Total', 0)
        total_aplicacion = totales.get('Aplicacion', {}).get('Total', 0)
        diferencia = total_origen - total_aplicacion
        
        ws.append([
            "Total Origen",
            "",
            "",
            total_origen
        ])
        ws.append([
            "Total Aplicación",
            "",
            "",
            total_aplicacion
        ])
        ws.append([
            "Diferencia",
            "",
            "",
            diferencia
        ])
        
        # Aplicar estilos
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.row <= 3 or (cell.value and isinstance(cell.value, str) and ('ORIGEN' in str(cell.value).upper() or 'APLICACIÓN' in str(cell.value).upper() or 'TOTALES' in str(cell.value).upper())):
                    if cell.row <= 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.alignment = center_alignment
                elif cell.column in [2, 3, 4] and isinstance(cell.value, (int, float)):
                    cell.alignment = right_alignment
                    cell.number_format = '#,##0.00'
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        return wb
    except Exception as e:
        print(f"Error al exportar Origen y Aplicación: {e}")
        return None

def exportar_analisis_excel(anio_seleccionado=None, tipo_analisis=None, **kwargs):
    """Exporta el análisis financiero especificado a un archivo Excel
    
    Args:
        anio_seleccionado: Año del análisis
        tipo_analisis: 'vertical', 'horizontal', 'ratios', 'origen_aplicacion'
        **kwargs: Datos adicionales según el tipo de análisis
            - Para 'vertical': report_data
            - Para 'horizontal': analisis_comparativo, periodo_base, periodo_analisis
            - Para 'ratios': ratios_data
            - Para 'origen_aplicacion': origen_aplicacion_data, periodo_base, periodo_analisis
    """
    try:
        if tipo_analisis == 'vertical':
            report_data = kwargs.get('report_data')
            if not report_data:
                return None
            return exportar_analisis_vertical_excel(anio_seleccionado, report_data)
        
        elif tipo_analisis == 'horizontal':
            analisis_comparativo = kwargs.get('analisis_comparativo')
            periodo_base = kwargs.get('periodo_base')
            periodo_analisis = kwargs.get('periodo_analisis')
            if not analisis_comparativo:
                return None
            return exportar_analisis_horizontal_excel(periodo_base, periodo_analisis, analisis_comparativo)
        
        elif tipo_analisis == 'ratios':
            ratios_data = kwargs.get('ratios_data')
            if not ratios_data:
                return None
            return exportar_ratios_excel(anio_seleccionado, ratios_data)
        
        elif tipo_analisis == 'origen_aplicacion':
            origen_aplicacion_data = kwargs.get('origen_aplicacion_data')
            periodo_base = kwargs.get('periodo_base')
            periodo_analisis = kwargs.get('periodo_analisis')
            if not origen_aplicacion_data:
                return None
            return exportar_origen_aplicacion_excel(periodo_base, periodo_analisis, origen_aplicacion_data)
        
        # Si no se especifica tipo, retornar None (no exportar todos)
        return None
        
    except Exception as e:
        print(f"Error al exportar análisis a Excel: {e}")
        import traceback
        traceback.print_exc()
        return None


def calcular_ctno(anio_seleccionado):
    """
    Calcula el Capital de Trabajo Neto Operativo (CTNO) para un año específico.
    
    Fórmula: CTNO = (Total de Cuentas por Cobrar + Total de Inventarios) - Total de Cuentas por Pagar
    
    La función busca estas cuentas en el catálogo de cuentas por nombre:
    - Cuentas por Cobrar: busca cuentas con "por cobrar" o "cuentas por cobrar" en el nombre
    - Inventarios: busca cuentas con "inventario" en el nombre
    - Cuentas por Pagar: busca cuentas con "por pagar" o "cuentas por pagar" en el nombre
    
    Args:
        anio_seleccionado (int): Año para el cual calcular el CTNO
        
    Returns:
        dict: Diccionario con:
            - 'ctno': Valor del CTNO calculado
            - 'cuentas_por_cobrar': Total de cuentas por cobrar
            - 'inventarios': Total de inventarios
            - 'cuentas_por_pagar': Total de cuentas por pagar
            - 'anio': Año seleccionado
            - 'exito': True si se calculó correctamente, False en caso de error
    """
    try:
        with engine.connect() as conn:
            # Obtener el PeriodoID para el año seleccionado
            periodo_query = text("SELECT PeriodoID FROM Periodo WHERE Anio = :anio")
            periodo_result = conn.execute(periodo_query, {"anio": anio_seleccionado}).fetchone()
            
            if not periodo_result:
                return {
                    'ctno': 0.0,
                    'cuentas_por_cobrar': 0.0,
                    'inventarios': 0.0,
                    'cuentas_por_pagar': 0.0,
                    'anio': anio_seleccionado,
                    'exito': False,
                    'mensaje': f'No se encontró el período para el año {anio_seleccionado}'
                }
            
            periodo_id = periodo_result[0]
            
            # 1. Calcular Total de Cuentas por Cobrar
            # Primero busca cuentas que contengan "por cobrar" o "cuentas por cobrar" en el nombre
            cuentas_por_cobrar_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por cobrar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por cobrar%'
                    OR LOWER(c.NombreCuenta) LIKE '%cobrar%'
                )
                AND c.TipoCuenta = 'Activo'
            """)
            cuentas_por_cobrar_result = conn.execute(cuentas_por_cobrar_query, {"periodo_id": periodo_id}).fetchone()
            total_cuentas_por_cobrar = float(cuentas_por_cobrar_result[0]) if cuentas_por_cobrar_result and cuentas_por_cobrar_result[0] else 0.0
            
            # Si no hay cuentas por cobrar, buscar cuentas con "cliente" en el nombre
            if total_cuentas_por_cobrar == 0.0:
                clientes_query = text("""
                    SELECT COALESCE(SUM(s.Monto), 0) AS total
                    FROM CatalogoCuentas c
                    INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                    WHERE s.PeriodoID = :periodo_id
                    AND (
                        LOWER(c.NombreCuenta) LIKE '%cliente%'
                        OR LOWER(c.NombreCuenta) LIKE '%clientes%'
                    )
                    AND c.TipoCuenta = 'Activo'
                """)
                clientes_result = conn.execute(clientes_query, {"periodo_id": periodo_id}).fetchone()
                total_cuentas_por_cobrar = float(clientes_result[0]) if clientes_result and clientes_result[0] else 0.0
            
            # 2. Calcular Total de Inventarios
            # Busca cuentas que contengan "inventario" en el nombre
            inventarios_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id
                AND LOWER(c.NombreCuenta) LIKE '%inventario%'
                AND c.TipoCuenta = 'Activo'
            """)
            inventarios_result = conn.execute(inventarios_query, {"periodo_id": periodo_id}).fetchone()
            total_inventarios = float(inventarios_result[0]) if inventarios_result and inventarios_result[0] else 0.0
            
            # 3. Calcular Total de Cuentas por Pagar
            # Busca cuentas que contengan "por pagar" o "cuentas por pagar" en el nombre
            cuentas_por_pagar_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por pagar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por pagar%'
                    OR LOWER(c.NombreCuenta) LIKE '%pagar%'
                )
                AND c.TipoCuenta = 'Pasivo'
            """)
            cuentas_por_pagar_result = conn.execute(cuentas_por_pagar_query, {"periodo_id": periodo_id}).fetchone()
            total_cuentas_por_pagar = float(cuentas_por_pagar_result[0]) if cuentas_por_pagar_result and cuentas_por_pagar_result[0] else 0.0
            
            # 4. Calcular CTNO
            # CTNO = (Cuentas por Cobrar + Inventarios) - Cuentas por Pagar
            ctno = (total_cuentas_por_cobrar + total_inventarios) - total_cuentas_por_pagar
            
            return {
                'ctno': ctno,
                'cuentas_por_cobrar': total_cuentas_por_cobrar,
                'inventarios': total_inventarios,
                'cuentas_por_pagar': total_cuentas_por_pagar,
                'anio': anio_seleccionado,
                'exito': True,
                'mensaje': 'CTNO calculado correctamente'
            }

    except Exception as e:
        print(f"Error al calcular CTNO: {e}")
        import traceback
        traceback.print_exc()
        return {
            'ctno': 0.0,
            'cuentas_por_cobrar': 0.0,
            'inventarios': 0.0,
            'cuentas_por_pagar': 0.0,
            'anio': anio_seleccionado,
            'exito': False,
            'mensaje': f'Error al calcular CTNO: {str(e)}'
        }

def calcular_feo_indirecto(fecha_inicio, fecha_fin):
    """
    Calcula el Flujo de Efectivo Operativo (FEO) usando el Método Indirecto.
    
    Fórmula: FEO = Utilidad Neta + Gastos no monetarios - Cambios en Capital de Trabajo Operativo
    
    Args:
        fecha_inicio (str o date): Fecha de inicio del período (formato 'YYYY-MM-DD' o datetime.date)
        fecha_fin (str o date): Fecha de fin del período (formato 'YYYY-MM-DD' o datetime.date)
        
    Returns:
        dict: Diccionario con:
            - 'feo': Valor del FEO calculado
            - 'utilidad_neta': Utilidad neta del período
            - 'gastos_no_monetarios': Suma de depreciación y amortización
            - 'cambio_cuentas_por_cobrar': Cambio en cuentas por cobrar
            - 'cambio_inventarios': Cambio en inventarios
            - 'cambio_cuentas_por_pagar': Cambio en cuentas por pagar
            - 'fecha_inicio': Fecha de inicio
            - 'fecha_fin': Fecha de fin
            - 'exito': True si se calculó correctamente, False en caso de error
            - 'mensaje': Mensaje descriptivo del resultado
    """
    try:
        from datetime import datetime, date
        
        # Convertir fechas a objetos date si son strings
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        if isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Obtener los años de inicio y fin
        anio_inicio = fecha_inicio.year
        anio_fin = fecha_fin.year
        
        with engine.connect() as conn:
            # Obtener PeriodoID para el año de inicio y fin
            periodo_inicio_query = text("SELECT PeriodoID FROM Periodo WHERE Anio = :anio")
            periodo_inicio_result = conn.execute(periodo_inicio_query, {"anio": anio_inicio}).fetchone()
            periodo_fin_result = conn.execute(periodo_inicio_query, {"anio": anio_fin}).fetchone()
            
            if not periodo_inicio_result or not periodo_fin_result:
                return {
                    'feo': 0.0,
                    'utilidad_neta': 0.0,
                    'gastos_no_monetarios': 0.0,
                    'cambio_cuentas_por_cobrar': 0.0,
                    'cambio_inventarios': 0.0,
                    'cambio_cuentas_por_pagar': 0.0,
                    'fecha_inicio': str(fecha_inicio),
                    'fecha_fin': str(fecha_fin),
                    'exito': False,
                    'mensaje': f'No se encontraron períodos para los años {anio_inicio} y/o {anio_fin}'
                }
            
            periodo_id_inicio = periodo_inicio_result[0]
            periodo_id_fin = periodo_fin_result[0]
            
            # 1. Calcular Utilidad Neta
            # Ingresos Totales
            ingresos_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND c.TipoCuenta = 'Ingreso'
            """)
            ingresos_result = conn.execute(ingresos_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            ingresos_totales = float(ingresos_result[0]) if ingresos_result and ingresos_result[0] else 0.0
            
            # Costos Totales
            costos_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND c.TipoCuenta = 'Costo'
            """)
            costos_result = conn.execute(costos_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            costos_totales = float(costos_result[0]) if costos_result and costos_result[0] else 0.0
            
            # Gastos Totales
            gastos_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND c.TipoCuenta = 'Gasto'
            """)
            gastos_result = conn.execute(gastos_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            gastos_totales = float(gastos_result[0]) if gastos_result and gastos_result[0] else 0.0
            
            # Utilidad Neta = Ingresos - Costos - Gastos
            utilidad_neta = ingresos_totales - costos_totales - gastos_totales
            
            # 2. Sumar Gastos No Monetarios (Depreciación y Amortización)
            gastos_no_monetarios_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND c.TipoCuenta = 'Gasto'
                AND (
                    LOWER(c.NombreCuenta) LIKE '%depreciaci%'
                    OR LOWER(c.NombreCuenta) LIKE '%amortizaci%'
                    OR LOWER(c.NombreCuenta) LIKE '%deprecioaci%'
                )
            """)
            gastos_no_monetarios_result = conn.execute(gastos_no_monetarios_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            gastos_no_monetarios = abs(float(gastos_no_monetarios_result[0])) if gastos_no_monetarios_result and gastos_no_monetarios_result[0] else 0.0
            
            # Utilidad Ajustada = Utilidad Neta + Gastos No Monetarios
            utilidad_ajustada = utilidad_neta + gastos_no_monetarios
            
            # 3. Calcular Cambios en Capital de Trabajo
            
            # 3.1 Cambio en Cuentas por Cobrar
            # Saldo en fecha_fin
            cxc_fin_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por cobrar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por cobrar%'
                    OR LOWER(c.NombreCuenta) LIKE '%cobrar%'
                )
                AND c.TipoCuenta = 'Activo'
            """)
            cxc_fin_result = conn.execute(cxc_fin_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            cxc_fin = float(cxc_fin_result[0]) if cxc_fin_result and cxc_fin_result[0] else 0.0
            
            # Si no hay cuentas por cobrar, buscar clientes
            if cxc_fin == 0.0:
                clientes_fin_query = text("""
                    SELECT COALESCE(SUM(s.Monto), 0) AS total
                    FROM CatalogoCuentas c
                    INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                    WHERE s.PeriodoID = :periodo_id_fin
                    AND (
                        LOWER(c.NombreCuenta) LIKE '%cliente%'
                        OR LOWER(c.NombreCuenta) LIKE '%clientes%'
                    )
                    AND c.TipoCuenta = 'Activo'
                """)
                clientes_fin_result = conn.execute(clientes_fin_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
                cxc_fin = float(clientes_fin_result[0]) if clientes_fin_result and clientes_fin_result[0] else 0.0
            
            # Saldo en fecha_inicio
            cxc_inicio_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_inicio
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por cobrar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por cobrar%'
                    OR LOWER(c.NombreCuenta) LIKE '%cobrar%'
                )
                AND c.TipoCuenta = 'Activo'
            """)
            cxc_inicio_result = conn.execute(cxc_inicio_query, {"periodo_id_inicio": periodo_id_inicio}).fetchone()
            cxc_inicio = float(cxc_inicio_result[0]) if cxc_inicio_result and cxc_inicio_result[0] else 0.0
            
            # Si no hay cuentas por cobrar, buscar clientes
            if cxc_inicio == 0.0:
                clientes_inicio_query = text("""
                    SELECT COALESCE(SUM(s.Monto), 0) AS total
                    FROM CatalogoCuentas c
                    INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                    WHERE s.PeriodoID = :periodo_id_inicio
                    AND (
                        LOWER(c.NombreCuenta) LIKE '%cliente%'
                        OR LOWER(c.NombreCuenta) LIKE '%clientes%'
                    )
                    AND c.TipoCuenta = 'Activo'
                """)
                clientes_inicio_result = conn.execute(clientes_inicio_query, {"periodo_id_inicio": periodo_id_inicio}).fetchone()
                cxc_inicio = float(clientes_inicio_result[0]) if clientes_inicio_result and clientes_inicio_result[0] else 0.0
            
            # Cambio en Cuentas por Cobrar = Saldo fin - Saldo inicio
            cambio_cxc = cxc_fin - cxc_inicio
            
            # 3.2 Cambio en Inventarios
            # Saldo en fecha_fin
            inventarios_fin_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND LOWER(c.NombreCuenta) LIKE '%inventario%'
                AND c.TipoCuenta = 'Activo'
            """)
            inventarios_fin_result = conn.execute(inventarios_fin_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            inventarios_fin = float(inventarios_fin_result[0]) if inventarios_fin_result and inventarios_fin_result[0] else 0.0
            
            # Saldo en fecha_inicio
            inventarios_inicio_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_inicio
                AND LOWER(c.NombreCuenta) LIKE '%inventario%'
                AND c.TipoCuenta = 'Activo'
            """)
            inventarios_inicio_result = conn.execute(inventarios_inicio_query, {"periodo_id_inicio": periodo_id_inicio}).fetchone()
            inventarios_inicio = float(inventarios_inicio_result[0]) if inventarios_inicio_result and inventarios_inicio_result[0] else 0.0
            
            # Cambio en Inventarios = Saldo fin - Saldo inicio
            cambio_inventarios = inventarios_fin - inventarios_inicio
            
            # 3.3 Cambio en Cuentas por Pagar
            # Saldo en fecha_fin
            cxp_fin_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_fin
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por pagar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por pagar%'
                    OR LOWER(c.NombreCuenta) LIKE '%pagar%'
                )
                AND c.TipoCuenta = 'Pasivo'
            """)
            cxp_fin_result = conn.execute(cxp_fin_query, {"periodo_id_fin": periodo_id_fin}).fetchone()
            cxp_fin = float(cxp_fin_result[0]) if cxp_fin_result and cxp_fin_result[0] else 0.0
            
            # Saldo en fecha_inicio
            cxp_inicio_query = text("""
                SELECT COALESCE(SUM(s.Monto), 0) AS total
                FROM CatalogoCuentas c
                INNER JOIN SaldoCuenta s ON s.CuentaID = c.CuentaID
                WHERE s.PeriodoID = :periodo_id_inicio
                AND (
                    LOWER(c.NombreCuenta) LIKE '%por pagar%' 
                    OR LOWER(c.NombreCuenta) LIKE '%cuentas por pagar%'
                    OR LOWER(c.NombreCuenta) LIKE '%pagar%'
                )
                AND c.TipoCuenta = 'Pasivo'
            """)
            cxp_inicio_result = conn.execute(cxp_inicio_query, {"periodo_id_inicio": periodo_id_inicio}).fetchone()
            cxp_inicio = float(cxp_inicio_result[0]) if cxp_inicio_result and cxp_inicio_result[0] else 0.0
            
            # Cambio en Cuentas por Pagar = Saldo fin - Saldo inicio
            cambio_cxp = cxp_fin - cxp_inicio
            
            # 4. Calcular FEO Final
            # FEO = Utilidad Ajustada - Cambio CxC - Cambio Inventarios + Cambio CxP
            feo = utilidad_ajustada - cambio_cxc - cambio_inventarios + cambio_cxp
            
            return {
                'feo': feo,
                'utilidad_neta': utilidad_neta,
                'ingresos_totales': ingresos_totales,
                'costos_totales': costos_totales,
                'gastos_totales': gastos_totales,
                'gastos_no_monetarios': gastos_no_monetarios,
                'utilidad_ajustada': utilidad_ajustada,
                'cambio_cuentas_por_cobrar': cambio_cxc,
                'cxc_inicio': cxc_inicio,
                'cxc_fin': cxc_fin,
                'cambio_inventarios': cambio_inventarios,
                'inventarios_inicio': inventarios_inicio,
                'inventarios_fin': inventarios_fin,
                'cambio_cuentas_por_pagar': cambio_cxp,
                'cxp_inicio': cxp_inicio,
                'cxp_fin': cxp_fin,
                'fecha_inicio': str(fecha_inicio),
                'fecha_fin': str(fecha_fin),
                'anio_inicio': anio_inicio,
                'anio_fin': anio_fin,
                'exito': True,
                'mensaje': 'FEO calculado correctamente'
            }
            
    except Exception as e:
        print(f"Error al calcular FEO: {e}")
        import traceback
        traceback.print_exc()
        return {
            'feo': 0.0,
            'utilidad_neta': 0.0,
            'gastos_no_monetarios': 0.0,
            'cambio_cuentas_por_cobrar': 0.0,
            'cambio_inventarios': 0.0,
            'cambio_cuentas_por_pagar': 0.0,
            'fecha_inicio': str(fecha_inicio) if 'fecha_inicio' in locals() else '',
            'fecha_fin': str(fecha_fin) if 'fecha_fin' in locals() else '',
            'exito': False,
            'mensaje': f'Error al calcular FEO: {str(e)}'
        }

def calcular_estado_flujo_efectivo(periodo_inicio, periodo_fin):
    """
    Calcula el Estado de Flujo de Efectivo usando el Método Indirecto.
    Wrapper para la clase CashFlowEngine.
    """
    try:
        report_ant = get_financial_reports(periodo_inicio)
        report_act = get_financial_reports(periodo_fin)
        
        if not report_ant or not report_act:
             return {
                 'exito': False, 
                 'mensaje': f'No se encontraron datos para los periodos {periodo_inicio} y {periodo_fin}.',
                 'Operacion': {'detalles': [], 'total': 0.0},
                 'Inversion': {'detalles': [], 'total': 0.0},
                 'Financiamiento': {'detalles': [], 'total': 0.0},
                 'Validacion': {}
             }

        # Instanciar y ejecutar el motor de cálculo
        engine_fe = CashFlowEngine(report_ant, report_act, report_act)
        flujo = engine_fe.ejecutar()
        
        # Agregar metadatos para la vista
        flujo['periodo_inicio'] = periodo_inicio
        flujo['periodo_fin'] = periodo_fin
        flujo['exito'] = True
        
        return flujo
        
    except Exception as e:
        print(f"Error en CashFlowEngine: {e}")
        import traceback
        traceback.print_exc()
        return {
            'exito': False, 
            'mensaje': str(e),
            'Operacion': {'detalles': [], 'total': 0.0},
            'Inversion': {'detalles': [], 'total': 0.0},
            'Financiamiento': {'detalles': [], 'total': 0.0},
            'Validacion': {}
        }

class CashFlowEngine:
    def __init__(self, balance_anterior, balance_actual, estado_resultados):
        self.bg_ant = balance_anterior
        self.bg_act = balance_actual
        self.er_act = estado_resultados
        self.flujo = {
            'Operacion': {'detalles': [], 'total': 0.0},
            'Inversion': {'detalles': [], 'total': 0.0},
            'Financiamiento': {'detalles': [], 'total': 0.0},
            'Validacion': {}
        }
        self.depreciacion = 0.0
        self.utilidad_neta = 0.0

    def ejecutar(self):
        # 1. Obtener datos base
        self.utilidad_neta = self.er_act['Totales'].get('Utilidad Neta', 0.0)
        self.obtener_depreciacion()
        
        # 2. Calcular secciones
        self.calcular_actividades_operacion()
        self.calcular_actividades_inversion()
        self.calcular_actividades_financiamiento()
        
        # 3. Validar
        self.validar_flujo()
        
        return self.flujo

    def obtener_depreciacion(self):
        # BUG FIX: Calcular variación de Depreciación Acumulada (Balance General)
        # Fórmula: Depreciación del Periodo = Depreciación Acumulada Actual - Depreciación Acumulada Anterior
        
        def get_accumulated_depreciation(report):
            total = 0.0
            # Buscar en todo el Activo (usando el helper existente)
            cuentas = self._flatten_report_section(report, 'Activo')
            for c in cuentas.values():
                nombre = c['nombre'].lower()
                if 'depreciaci' in nombre or 'amortizaci' in nombre:
                    # Depreciación acumulada suele ser negativa (contra-activo). Usamos abs.
                    total += abs(c['monto'])
            return total

        dep_acum_act = get_accumulated_depreciation(self.bg_act)
        dep_acum_ant = get_accumulated_depreciation(self.bg_ant)
        
        # La diferencia es el cargo del periodo
        self.depreciacion = dep_acum_act - dep_acum_ant

    def calcular_actividades_operacion(self):
        total_op = 0.0
        detalles = []

        # A. Utilidad Neta
        detalles.append({'concepto': 'Utilidad Neta', 'monto': self.utilidad_neta})
        total_op += self.utilidad_neta

        # B. Partidas virtuales (Depreciación) - REGLA DE ORO 1
        # Se agrega forzosamente si hay depreciación detectada
        if self.depreciacion > 0:
            detalles.append({'concepto': 'Más: Cargos por Depreciación y Amortización', 'monto': self.depreciacion})
            total_op += self.depreciacion

        # C. Cambios en Capital de Trabajo (Activos/Pasivos Operativos)
        # REGLA 3A: Clientes, Inventarios, Anticipos, Impuestos, Proveedores, etc.
        
        # --- ACTIVOS OPERATIVOS ---
        # Recorremos TODO el activo para buscar cuentas operativas
        activos_ant = self._flatten_report_section(self.bg_ant, 'Activo')
        activos_act = self._flatten_report_section(self.bg_act, 'Activo')
        todas_cuentas_act = set(activos_ant.keys()) | set(activos_act.keys())

        for cid in todas_cuentas_act:
            nombre = activos_act.get(cid, {}).get('nombre') or activos_ant.get(cid, {}).get('nombre')
            if not nombre: continue
            
            # Filtrar solo operativos
            if self._es_operativo_activo(nombre):
                saldo_ant = activos_ant.get(cid, {}).get('monto', 0.0)
                saldo_act = activos_act.get(cid, {}).get('monto', 0.0)
                
                # Regla Activos: Anterior - Actual
                variacion = saldo_ant - saldo_act
                
                if abs(variacion) > 0.01:
                    detalles.append({'concepto': f'Cambio en {nombre}', 'monto': variacion})
                    total_op += variacion

        # --- PASIVOS OPERATIVOS ---
        pasivos_ant = self._flatten_report_section(self.bg_ant, 'Pasivo')
        pasivos_act = self._flatten_report_section(self.bg_act, 'Pasivo')
        todas_cuentas_pas = set(pasivos_ant.keys()) | set(pasivos_act.keys())

        for cid in todas_cuentas_pas:
            nombre = pasivos_act.get(cid, {}).get('nombre') or pasivos_ant.get(cid, {}).get('nombre')
            if not nombre: continue

            # Filtrar solo operativos
            if self._es_operativo_pasivo(nombre):
                saldo_ant = pasivos_ant.get(cid, {}).get('monto', 0.0)
                saldo_act = pasivos_act.get(cid, {}).get('monto', 0.0)
                
                # Regla Pasivos: Actual - Anterior
                variacion = saldo_act - saldo_ant
                
                if abs(variacion) > 0.01:
                    detalles.append({'concepto': f'Cambio en {nombre}', 'monto': variacion})
                    total_op += variacion

        self.flujo['Operacion']['detalles'] = detalles
        self.flujo['Operacion']['total'] = total_op

    def calcular_actividades_inversion(self):
        total_inv = 0.0
        detalles = []

        # REGLA 2 y 3B: Solo Activos Fijos Reales (Maquinaria, Edificios, etc.)
        # IGNORAR Depreciación Acumulada
        
        activos_ant = self._flatten_report_section(self.bg_ant, 'Activo')
        activos_act = self._flatten_report_section(self.bg_act, 'Activo')
        todas_cuentas = set(activos_ant.keys()) | set(activos_act.keys())
        
        for cid in todas_cuentas:
            nombre = activos_act.get(cid, {}).get('nombre') or activos_ant.get(cid, {}).get('nombre')
            if not nombre: continue

            if self._es_inversion(nombre):
                saldo_ant = activos_ant.get(cid, {}).get('monto', 0.0)
                saldo_act = activos_act.get(cid, {}).get('monto', 0.0)
                
                # Regla Inversión: Variación = Saldo Actual - Saldo Anterior
                # Flujo = -(Variación)  (Aumento de activo es salida de dinero)
                variacion = saldo_act - saldo_ant
                flujo = -variacion
                
                if abs(flujo) > 0.01:
                    detalles.append({'concepto': f'Adquisición/Venta de {nombre}', 'monto': flujo})
                    total_inv += flujo

        self.flujo['Inversion']['detalles'] = detalles
        self.flujo['Inversion']['total'] = total_inv

    def calcular_actividades_financiamiento(self):
        total_fin = 0.0
        detalles = []

        # REGLA 3C: Capital Social, Préstamos
        
        # 1. Pasivos Financieros
        pasivos_ant = self._flatten_report_section(self.bg_ant, 'Pasivo')
        pasivos_act = self._flatten_report_section(self.bg_act, 'Pasivo')
        todas_cuentas_pas = set(pasivos_ant.keys()) | set(pasivos_act.keys())

        for cid in todas_cuentas_pas:
            nombre = pasivos_act.get(cid, {}).get('nombre') or pasivos_ant.get(cid, {}).get('nombre')
            if not nombre: continue

            if self._es_financiamiento_pasivo(nombre):
                saldo_ant = pasivos_ant.get(cid, {}).get('monto', 0.0)
                saldo_act = pasivos_act.get(cid, {}).get('monto', 0.0)
                
                # Regla Pasivos: Actual - Anterior
                variacion = saldo_act - saldo_ant
                
                if abs(variacion) > 0.01:
                    detalles.append({'concepto': f'Variación en {nombre}', 'monto': variacion})
                    total_fin += variacion

        # 2. Patrimonio (Capital Social)
        patrimonio_ant = self._flatten_report_section(self.bg_ant, 'Patrimonio')
        patrimonio_act = self._flatten_report_section(self.bg_act, 'Patrimonio')
        todas_cuentas_pat = set(patrimonio_ant.keys()) | set(patrimonio_act.keys())
        
        utilidad_acumulada_ant = 0.0
        utilidad_acumulada_act = 0.0
        
        for cid in todas_cuentas_pat:
            nombre = patrimonio_act.get(cid, {}).get('nombre') or patrimonio_ant.get(cid, {}).get('nombre')
            if not nombre: continue
            nombre_lower = nombre.lower()
            
            saldo_ant = patrimonio_ant.get(cid, {}).get('monto', 0.0)
            saldo_act = patrimonio_act.get(cid, {}).get('monto', 0.0)
            
            # Identificar Utilidades para el cálculo de dividendos
            if any(x in nombre_lower for x in ['utilidad', 'resultado', 'ganancia', 'perdida', 'ejercicio', 'acumulada']):
                utilidad_acumulada_ant += saldo_ant
                utilidad_acumulada_act += saldo_act
                continue 
            
            # Otras cuentas de capital
            variacion = saldo_act - saldo_ant
            if abs(variacion) > 0.01:
                detalles.append({'concepto': f'Variación en {nombre}', 'monto': variacion})
                total_fin += variacion
        
        # 3. Ajuste de Patrimonio (Dividendos)
        # Fórmula: (Utilidad Retenida Actual - Utilidad Retenida Anterior) - Utilidad Neta Actual
        # Si es negativo, significa que salieron dividendos.
        
        cambio_utilidades_retenidas = utilidad_acumulada_act - utilidad_acumulada_ant
        diferencia = cambio_utilidades_retenidas - self.utilidad_neta
        
        if abs(diferencia) > 0.01:
            concepto = 'Pago de Dividendos / Ajustes a Resultados'
            detalles.append({'concepto': concepto, 'monto': diferencia})
            total_fin += diferencia

        self.flujo['Financiamiento']['detalles'] = detalles
        self.flujo['Financiamiento']['total'] = total_fin

    def validar_flujo(self):
        flujo_neto = (self.flujo['Operacion']['total'] + 
                      self.flujo['Inversion']['total'] + 
                      self.flujo['Financiamiento']['total'])
        
        efectivo_ini = self._calcular_efectivo_total(self.bg_ant)
        efectivo_fin = self._calcular_efectivo_total(self.bg_act)
        
        efectivo_calculado = efectivo_ini + flujo_neto
        diferencia = efectivo_fin - efectivo_calculado
        
        self.flujo['Validacion'] = {
            'flujo_neto': flujo_neto,
            'efectivo_inicial': efectivo_ini,
            'efectivo_final_real': efectivo_fin,
            'efectivo_final_calculado': efectivo_calculado,
            'diferencia': diferencia,
            'cuadra': abs(diferencia) < 1.0
        }

    # --- HELPER FUNCTIONS ---

    def _flatten_report_section(self, reporte, seccion):
        """Aplana la estructura jerárquica de una sección del reporte"""
        cuentas = {}
        if seccion in reporte:
            for subtipo, lista in reporte[seccion].items():
                for c in lista:
                    cuentas[c['id']] = c
        return cuentas

    def _es_efectivo(self, nombre):
        if not nombre: return False
        n = nombre.lower()
        return any(x in n for x in ['caja', 'banco', 'efectivo', 'cash', 'disponible'])

    def _es_operativo_activo(self, nombre):
        """Regla 3A: Clientes, Inventarios, Anticipos, etc."""
        if not nombre: return False
        n = nombre.lower()
        
        # Excluir explícitamente efectivo e inversión
        if self._es_efectivo(nombre): return False
        if self._es_inversion(nombre): return False
        
        # Palabras clave operativas
        keywords = ['cliente', 'cobrar', 'inventario', 'almacen', 'anticipo', 
                   'deposito', 'garantia', 'otro activo', 'impuesto', 'renta', 'iva', 'acreditable',
                   'obra', 'trabajo', 'proceso', 'pago anticipado', 'pagos anticipados']
        
        return any(x in n for x in keywords)

    def _es_operativo_pasivo(self, nombre):
        """Regla 3A: Proveedores, Impuestos, etc."""
        if not nombre: return False
        n = nombre.lower()
        
        # Excluir financiamiento
        if self._es_financiamiento_pasivo(nombre): return False
        
        # Palabras clave operativas
        keywords = ['proveedor', 'acreedor', 'por pagar', 'impuesto', 'retencion', 
                   'iva', 'acumulado', 'laboral', 'sueldo', 'salario']
        
        return any(x in n for x in keywords)

    def _es_inversion(self, nombre):
        """Regla 3B: Activos Fijos Tangibles"""
        if not nombre: return False
        n = nombre.lower()
        
        # EXCLUIR DEPRECIACIÓN (Regla de Oro 2)
        if 'depreciaci' in n or 'amortizaci' in n: return False
        
        # Excluir Obras en Proceso (Regla Hotfix 1)
        if 'proceso' in n or 'obra' in n or 'trabajo' in n: return False
        
        keywords = ['maquinaria', 'equipo', 'edificio', 'terreno', 'vehiculo', 
                   'rodante', 'mobiliario', 'construccion', 'propiedad']
        
        return any(x in n for x in keywords)

    def _es_financiamiento_pasivo(self, nombre):
        """Regla 3C: Préstamos Bancarios"""
        if not nombre: return False
        n = nombre.lower()
        
        keywords = ['prestamo', 'préstamo', 'credito', 'crédito', 'bancari', 'financier', 'hipoteca']
        
        return any(x in n for x in keywords)

    def _calcular_efectivo_total(self, reporte):
        total = 0.0
        # Buscar en todo el activo, no solo corriente, por si acaso
        activos = self._flatten_report_section(reporte, 'Activo')
        for c in activos.values():
            if self._es_efectivo(c['nombre']):
                total += c['monto']
        return total

def generar_analisis_dupont(anio_actual):
    """
    Calcula el análisis DuPont de 3 factores para el año actual y el anterior.
    Retorna KPIs de Margen Neto, Rotación de Activos y Multiplicador de Capital.
    """
    try:
        anio_anterior = anio_actual - 1
        
        report_act = get_financial_reports(anio_actual)
        report_ant = get_financial_reports(anio_anterior)
        
        # Si no hay datos del año anterior, intentamos calcular solo el actual
        # pero para la comparativa necesitamos ambos.
        if not report_act:
             return {'exito': False, 'mensaje': f'No se encontraron datos para el año {anio_actual}.'}
        
        # Si falta el anterior, creamos un dummy o retornamos error parcial?
        # El usuario pide comparativa. Asumiremos que existen o manejaremos ceros.
        if not report_ant:
            report_ant = {'Totales': {}} # Dummy vacío

        def calcular_kpis(report):
            totales = report.get('Totales', {})
            utilidad_neta = totales.get('Utilidad Neta', 0.0)
            ventas = totales.get('Ingreso', 0.0)
            activos = totales.get('Activo', 0.0)
            patrimonio = totales.get('Patrimonio', 0.0)
            
            # 1. Margen de Utilidad Neta (Eficiencia)
            margen_neto = (utilidad_neta / ventas) if ventas else 0.0
            
            # 2. Rotación de Activos Totales (Eficiencia de Activos)
            rotacion_activos = (ventas / activos) if activos else 0.0
            
            # 3. Multiplicador de Capital (Apalancamiento)
            multiplicador = (activos / patrimonio) if patrimonio else 0.0
            
            # 4. ROE DuPont
            roe = margen_neto * rotacion_activos * multiplicador
            
            return {
                'margen_neto': margen_neto, # Se guarda como decimal (ej. 0.0624)
                'rotacion_activos': rotacion_activos,
                'multiplicador': multiplicador,
                'roe': roe
            }

        dupont_act = calcular_kpis(report_act)
        dupont_ant = calcular_kpis(report_ant)
        
        # Variaciones
        roe_act = dupont_act['roe']
        roe_ant = dupont_ant['roe']
        
        variacion_roe_decimal = 0.0
        if roe_ant:
            variacion_roe_decimal = (roe_act - roe_ant) / abs(roe_ant)
            cambio_roe_val = variacion_roe_decimal * 100
            cambio_roe = f"{cambio_roe_val:+.1f}%"
        else:
            cambio_roe = "N/A"
        
        # Factor determinante
        # Calculamos el impacto relativo o simplemente cuál varió más
        def calc_change(curr, prev):
            return ((curr - prev) / abs(prev)) if prev else 0.0
            
        cambio_margen = calc_change(dupont_act['margen_neto'], dupont_ant['margen_neto'])
        cambio_rotacion = calc_change(dupont_act['rotacion_activos'], dupont_ant['rotacion_activos'])
        cambio_multiplicador = calc_change(dupont_act['multiplicador'], dupont_ant['multiplicador'])
        
        factors = {
            'Margen de Utilidad': cambio_margen,
            'Rotación de Activos': cambio_rotacion,
            'Apalancamiento Financiero': cambio_multiplicador
        }
        
        # Determinamos cuál influyó más en la dirección del ROE
        # Si ROE bajó, buscamos el factor más negativo. Si subió, el más positivo.
        diff_roe = roe_act - roe_ant
        
        if diff_roe < 0:
            # El que más bajó (o menos subió)
            determinante = min(factors, key=factors.get)
        else:
            # El que más subió
            determinante = max(factors, key=factors.get)
            
        # Formatear texto del determinante
        if abs(factors[determinante]) < 0.01:
             determinante_texto = "Estable"
        else:
             direction = "disminución" if factors[determinante] < 0 else "aumento"
             determinante_texto = f"Impulsado por {direction} en {determinante}"

        return {
            'analisis_dupont': {
                str(anio_anterior): dupont_ant,
                str(anio_actual): dupont_act,
                'variaciones': {
                    'cambio_roe': cambio_roe,
                    'factor_determinante': determinante_texto,
                    'roe': variacion_roe_decimal
                }
            },
            'exito': True
        }
        
    except Exception as e:
        print(f"Error en DuPont: {e}")
        import traceback
        traceback.print_exc()
        return {'exito': False, 'mensaje': str(e)}

def generar_estado_proforma(estado_resultados_base, tasa_crecimiento):
    """
    Genera un Estado de Resultados Proforma proyectado basado en porcentaje de ventas.
    
    Args:
        estado_resultados_base (dict): Reporte financiero del año base.
        tasa_crecimiento (float): Tasa de crecimiento de ventas (ej. 0.15 para 15%).
        
    Returns:
        dict: Objeto con datos comparativos Real vs Proyectado.
    """
    try:
        totales = estado_resultados_base.get('Totales', {})
        
        # 1. Datos Base
        ventas_base = float(totales.get('Ingreso', 0.0))
        costos_base = float(totales.get('Costo', 0.0))
        gastos_base = float(totales.get('Gasto', 0.0))
        utilidad_neta_base = float(totales.get('Utilidad Neta', 0.0))
        
        # Calcular otros subtotales base
        utilidad_bruta_base = ventas_base - costos_base
        # Asumimos que Utilidad Operativa es Bruta - Gastos
        utilidad_operativa_base = utilidad_bruta_base - gastos_base
        
        # Impuestos base (inferido: Utilidad Operativa - Utilidad Neta)
        impuestos_base = utilidad_operativa_base - utilidad_neta_base
        
        # 2. Proyección
        
        # A. Ventas
        ventas_proy = ventas_base * (1 + tasa_crecimiento)
        
        # B. Costos (Variable % de ventas)
        pct_costo = (costos_base / ventas_base) if ventas_base else 0.0
        costos_proy = ventas_proy * pct_costo
        
        # C. Utilidad Bruta
        utilidad_bruta_proy = ventas_proy - costos_proy
        
        # D. Gastos Operativos (Variable % de ventas)
        pct_gasto = (gastos_base / ventas_base) if ventas_base else 0.0
        gastos_proy = ventas_proy * pct_gasto
        
        # E. Utilidad Operativa (Antes de Impuestos)
        utilidad_operativa_proy = utilidad_bruta_proy - gastos_proy
        
        # F. Impuestos - REGLA DEL 30% (Hardcoded Tax Rate)
        # No proyectamos el impuesto anterior. Usamos tasa fija del 30%.
        impuestos_proy = 0.0
        if utilidad_operativa_proy > 0:
            impuestos_proy = utilidad_operativa_proy * 0.30
            
        # G. Utilidad Neta
        utilidad_neta_proy = utilidad_operativa_proy - impuestos_proy
        
        # 3. Construir Salida
        return {
            "proforma": {
                "escenario": f"Proyección con crecimiento del {tasa_crecimiento*100:.1f}%",
                "datos": {
                    "ingresos": { 
                        "base": ventas_base, 
                        "proyectado": ventas_proy, 
                        "variacion": ventas_proy - ventas_base 
                    },
                    "costos": { 
                        "base": costos_base, 
                        "proyectado": costos_proy, 
                        "variacion": costos_proy - costos_base 
                    },
                    "utilidad_bruta": { 
                        "base": utilidad_bruta_base, 
                        "proyectado": utilidad_bruta_proy, 
                        "variacion": utilidad_bruta_proy - utilidad_bruta_base 
                    },
                    "gastos_operativos": { 
                        "base": gastos_base, 
                        "proyectado": gastos_proy, 
                        "variacion": gastos_proy - gastos_base 
                    },
                    "utilidad_antes_impuestos": {
                        "base": utilidad_operativa_base,
                        "proyectado": utilidad_operativa_proy,
                        "variacion": utilidad_operativa_proy - utilidad_operativa_base
                    },
                    "impuestos": {
                        "base": impuestos_base,
                        "proyectado": impuestos_proy,
                        "variacion": impuestos_proy - impuestos_base
                    },
                    "utilidad_neta": { 
                        "base": utilidad_neta_base, 
                        "proyectado": utilidad_neta_proy, 
                        "variacion": utilidad_neta_proy - utilidad_neta_base 
                    }
                }
            },
            "exito": True
        }
    except Exception as e:
        print(f"Error en Proforma: {e}")
        import traceback
        traceback.print_exc()
        return {'exito': False, 'mensaje': str(e)}
